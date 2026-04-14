"""
Views for Django Form Workflows

This module provides the core views for form submission, approval workflows,
and submission management.
"""

import json
import logging
import re
from datetime import date, datetime, time
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.db import models
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.html import escape
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.http import require_http_methods

from .forms import DynamicForm
from .models import ApprovalTask, AuditLog, FormDefinition, FormField, FormSubmission
from .utils import (
    check_rate_limit,
    user_can_approve,
    user_can_submit_form,
    user_can_view_form,
)

logger = logging.getLogger(__name__)


def _get_accessible_category_pks(user):
    """Return the set of FormCategory PKs the user is permitted to access.

    A category is accessible when the user satisfies the group restriction at
    **every** level of the ancestor chain:

    * If a category has its own ``allowed_groups``, the user must be in at
      least one of them.
    * If a category has *no* ``allowed_groups`` of its own, access is
      inherited from the parent (or granted freely if there is no parent).

    This means sub-categories without explicit groups automatically inherit
    their parent's restriction, so adding "Students - All" to Student Services
    automatically gates On-Campus and Online as well.
    """
    from .models import FormCategory

    user_group_ids = set(user.groups.values_list("id", flat=True))

    all_cats = list(
        FormCategory.objects.prefetch_related("allowed_groups").order_by("id")
    )
    cat_by_pk = {cat.pk: cat for cat in all_cats}
    # Map each category to its set of required group IDs
    cat_groups = {
        cat.pk: set(cat.allowed_groups.values_list("id", flat=True)) for cat in all_cats
    }

    _cache = {}  # pk -> bool

    def _can_access(cat_pk):
        if cat_pk in _cache:
            return _cache[cat_pk]
        cat = cat_by_pk[cat_pk]
        groups = cat_groups[cat_pk]
        if groups:
            # Has own restriction — user must be in at least one group
            result = bool(user_group_ids & groups)
        elif cat.parent_id is None:
            # Top-level with no restriction — open to all authenticated users
            result = True
        else:
            # Inherit from parent
            result = _can_access(cat.parent_id)
        _cache[cat_pk] = result
        return result

    for cat in all_cats:
        _can_access(cat.pk)

    return {pk for pk, ok in _cache.items() if ok}


def _build_grouped_forms(forms):
    """
    Return a hierarchical list of category-tree nodes for rendering grouped forms.

    Each node in the returned list is a dict::

        {
            "category": FormCategory | None,   # None = uncategorised bucket
            "forms":    [FormDefinition, ...], # forms directly under this category
            "children": [node, ...],           # child category nodes (same structure)
        }

    Only categories that contain at least one visible form (directly or via a
    descendant) are included.  Uncategorised forms are appended as the final
    node with ``"category": None``.

    The tree is built with a single call to ``FormCategory.objects.all()``
    plus the already-filtered ``forms`` queryset, so the number of SQL
    queries is small and constant regardless of nesting depth.
    """
    from .models import FormCategory

    # ---- 1. Bucket forms by category pk --------------------------------
    ordered = forms.order_by(
        "category__order",
        "category__name",
        "name",
    ).select_related("category")

    forms_by_cat = {}  # cat_pk -> [FormDefinition]
    uncategorised = []

    for form in ordered:
        if form.category_id is None:
            uncategorised.append(form)
        else:
            forms_by_cat.setdefault(form.category_id, []).append(form)

    if not forms_by_cat and not uncategorised:
        return []

    # ---- 2. Load all categories and build parent → children map --------
    all_cats = {
        cat.pk: cat for cat in FormCategory.objects.all().order_by("order", "name")
    }
    children_by_parent = {}  # parent_pk | None -> [FormCategory]
    for cat in all_cats.values():
        children_by_parent.setdefault(cat.parent_id, []).append(cat)

    # ---- 3. Determine which categories have visible forms (memoised) ----
    _has_forms_cache = {}

    def _subtree_has_forms(cat_pk):
        if cat_pk in _has_forms_cache:
            return _has_forms_cache[cat_pk]
        result = cat_pk in forms_by_cat
        if not result:
            for child in children_by_parent.get(cat_pk, []):
                if _subtree_has_forms(child.pk):
                    result = True
                    break
        _has_forms_cache[cat_pk] = result
        return result

    # ---- 4. Recursively assemble the tree ------------------------------
    def _build_subtree(cat):
        visible_children = [
            _build_subtree(child)
            for child in children_by_parent.get(cat.pk, [])
            if _subtree_has_forms(child.pk)
        ]
        return {
            "category": cat,
            "forms": forms_by_cat.get(cat.pk, []),
            "children": visible_children,
        }

    top_level = [
        _build_subtree(cat)
        for cat in children_by_parent.get(None, [])
        if _subtree_has_forms(cat.pk)
    ]

    if uncategorised:
        top_level.append({"category": None, "forms": uncategorised, "children": []})

    return top_level


def form_list(request):
    """List all active forms the current user has access to.

    For non-staff users two layers of group filtering are applied:

    1. **Form-level** — the user must be in ``submit_groups``, or the form
       has no ``submit_groups`` restriction.
    2. **Category-level** — the full ancestor chain is checked via
       :func:`_get_accessible_category_pks`.  A category with no
       ``allowed_groups`` of its own inherits the restriction of the nearest
       ancestor that has groups set.  A root category with no groups is open
       to all authenticated users.

    Staff / superusers bypass both filters and see every active form.

    Anonymous users only see forms with ``requires_login=False``.

    Context variables
    -----------------
    forms
        Raw queryset (for templates that just want a flat list).
    grouped_forms
        Ordered list of ``(FormCategory | None, [FormDefinition, ...])``.
    """
    if not request.user.is_authenticated:
        # Anonymous users — only public forms
        forms = FormDefinition.objects.filter(
            is_active=True, is_listed=True, requires_login=False
        ).select_related("category")
    elif request.user.is_staff or request.user.is_superuser:
        forms = FormDefinition.objects.filter(
            is_active=True, is_listed=True
        ).select_related("category")
    else:
        user_groups = request.user.groups.all()
        accessible_cat_pks = _get_accessible_category_pks(request.user)
        forms = (
            FormDefinition.objects.filter(is_active=True, is_listed=True)
            # Annotate group counts to distinguish "no restriction" from "restricted"
            .annotate(
                submit_group_count=models.Count("submit_groups", distinct=True),
                view_group_count=models.Count("view_groups", distinct=True),
            )
            # Form-level: user in view_groups OR form has no view restriction
            .filter(
                models.Q(view_groups__in=user_groups) | models.Q(view_group_count=0)
            )
            # Form-level: user in submit_groups OR form has no submit restriction
            .filter(
                models.Q(submit_groups__in=user_groups) | models.Q(submit_group_count=0)
            )
            # Category-level: no category, or category accessible via hierarchy
            # _get_accessible_category_pks walks the parent chain, so a child
            # category with no allowed_groups inherits its parent's restriction.
            .filter(
                models.Q(category__isnull=True)
                | models.Q(category_id__in=accessible_cat_pks)
            )
            .distinct()
            .select_related("category")
        )

    # --- Optional name search ---
    search_query = request.GET.get("q", "").strip()
    if search_query:
        forms = forms.filter(name__icontains=search_query)

    grouped_forms = _build_grouped_forms(forms)

    return render(
        request,
        "django_forms_workflows/form_list.html",
        {
            "forms": forms,
            "grouped_forms": grouped_forms,
            "search_query": search_query,
        },
    )


def form_qr_code(request, slug):
    """Return a QR-code image (SVG or PNG) encoding the form's submit URL.

    Query parameters
    ----------------
    format : ``svg`` (default) or ``png``
    size   : module scale factor for PNG (default 8)

    The view is intentionally public so QR images can be embedded in printed
    materials without authentication.
    """
    # Resolve (and 404-guard) the form before checking for the optional package
    # so that an inactive/missing slug always returns 404 regardless of whether
    # segno is installed.
    form_def = get_object_or_404(FormDefinition, slug=slug, is_active=True)

    try:
        import segno
    except ImportError:
        return HttpResponse(
            "QR code generation requires the 'segno' package. "
            "Install it with: pip install django-forms-workflows[qr]",
            status=501,
            content_type="text/plain",
        )

    submit_url = request.build_absolute_uri(
        reverse("forms_workflows:form_submit", args=[form_def.slug])
    )

    qr = segno.make(submit_url)

    fmt = request.GET.get("format", "svg").lower()
    if fmt == "png":
        import io

        scale = int(request.GET.get("size", 8))
        buf = io.BytesIO()
        qr.save(buf, kind="png", scale=scale, border=2)
        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type="image/png")
        response["Content-Disposition"] = f'inline; filename="{form_def.slug}-qr.png"'
        return response

    # Default: SVG
    import io

    buf = io.BytesIO()
    qr.save(buf, kind="svg", scale=4, border=2, svgclass="qr-code")
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="image/svg+xml")
    response["Content-Disposition"] = f'inline; filename="{form_def.slug}-qr.svg"'
    return response


def form_submit(request, slug):
    """Submit a form — supports both authenticated and anonymous (public) access."""
    form_def = get_object_or_404(FormDefinition, slug=slug, is_active=True)
    is_anonymous = not request.user.is_authenticated

    # ── Authentication gate ─────────────────────────────────────────────
    if is_anonymous:
        if form_def.requires_login:
            from django.contrib.auth.views import redirect_to_login

            return redirect_to_login(request.get_full_path())
    else:
        # Authenticated permission checks
        if not user_can_view_form(request.user, form_def):
            messages.error(request, "You don't have permission to access this form.")
            return redirect("forms_workflows:form_list")
        if not user_can_submit_form(request.user, form_def):
            messages.error(request, "You don't have permission to submit this form.")
            return redirect("forms_workflows:form_list")

    # ── Submission controls ─────────────────────────────────────────────
    if form_def.close_date and timezone.now() >= form_def.close_date:
        messages.error(request, "This form is no longer accepting submissions.")
        return redirect("forms_workflows:form_list")

    if form_def.max_submissions is not None:
        submitted_count = (
            FormSubmission.objects.filter(form_definition=form_def)
            .exclude(status="draft")
            .count()
        )
        if submitted_count >= form_def.max_submissions:
            messages.error(
                request, "This form has reached its maximum number of submissions."
            )
            return redirect("forms_workflows:form_list")

    if form_def.one_per_user and not is_anonymous:
        existing = (
            FormSubmission.objects.filter(
                form_definition=form_def, submitter=request.user
            )
            .exclude(status__in=["draft", "withdrawn"])
            .exists()
        )
        if existing:
            messages.error(request, "You have already submitted this form.")
            return redirect("forms_workflows:my_submissions")

    # Drafts are only available for authenticated users
    draft = None
    if not is_anonymous:
        draft = FormSubmission.objects.filter(
            form_definition=form_def, submitter=request.user, status="draft"
        ).first()

    user_or_none = request.user if not is_anonymous else None

    if request.method == "POST":
        # ── Rate limit anonymous submissions ────────────────────────────
        if is_anonymous and not check_rate_limit(request, slug):
            return render(
                request,
                "django_forms_workflows/rate_limited.html",
                {"form_def": form_def},
                status=429,
            )

        # ------------------------------------------------------------------ #
        # Draft save: only for authenticated users.                          #
        # ------------------------------------------------------------------ #
        if "save_draft" in request.POST and not is_anonymous:
            _skip = {"csrfmiddlewaretoken", "save_draft", "submit"}
            raw_data = {}
            for k in request.POST:
                if k in _skip:
                    continue
                values = request.POST.getlist(k)
                raw_data[k] = values[0] if len(values) == 1 else values

            draft_obj, created = FormSubmission.objects.update_or_create(
                form_definition=form_def,
                submitter=request.user,
                status="draft",
                defaults={
                    "form_data": raw_data,
                    "submission_ip": get_client_ip(request),
                    "user_agent": request.META.get("HTTP_USER_AGENT", ""),
                },
            )

            # Persist uploaded files with the draft
            if request.FILES:
                _save_draft_files(request.FILES, draft_obj, form_def)

            AuditLog.objects.create(
                action="update" if not created else "create",
                object_type="FormSubmission",
                object_id=draft_obj.id,
                user=request.user,
                user_ip=get_client_ip(request),
                comments="Saved as draft",
            )
            messages.success(request, "Draft saved successfully.")
            return redirect("forms_workflows:my_submissions")

        # ------------------------------------------------------------------ #
        # Full submission: validate then save.                                #
        # ------------------------------------------------------------------ #

        # Eagerly persist uploaded files so they survive a validation failure.
        stashed_files = (
            _stash_uploaded_files(request.FILES, form_def) if request.FILES else {}
        )

        form = DynamicForm(
            form_definition=form_def,
            user=user_or_none,
            data=request.POST,
            files=request.FILES,
            stashed_files=stashed_files or None,
        )

        if form.is_valid():
            submission = draft or FormSubmission(
                form_definition=form_def,
                submitter=user_or_none,
                submission_ip=get_client_ip(request),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )
            if not submission.pk:
                submission.form_data = {}
                submission.save()

            # Merge stashed files with any draft file data so files
            # uploaded on a previous attempt are preserved.
            _existing_files = dict(stashed_files)
            if draft and draft.form_data:
                for k, v in draft.form_data.items():
                    if isinstance(v, dict) and v.get("filename"):
                        _existing_files.setdefault(k, v)
                    elif isinstance(v, list) and v and isinstance(v[0], dict):
                        _existing_files.setdefault(k, v)

            submission.form_data = serialize_form_data(
                form.cleaned_data,
                submission_id=submission.pk,
                existing_file_data=_existing_files or None,
            )
            submission.form_data = _re_evaluate_calculated_fields(
                submission.form_data, form_def
            )

            # ── Payment gate ──────────────────────────────────────────
            if form_def.payment_enabled and form_def.payment_provider:
                submission.status = "pending_payment"
                submission.save()
                AuditLog.objects.create(
                    action="submit",
                    object_type="FormSubmission",
                    object_id=submission.id,
                    user=user_or_none,
                    user_ip=get_client_ip(request),
                    comments="Form validated, redirecting to payment",
                )
                return redirect(
                    "forms_workflows:payment_initiate",
                    submission_id=submission.id,
                )

            submission.status = "submitted"
            submission.submitted_at = timezone.now()
            submission.save()

            AuditLog.objects.create(
                action="submit",
                object_type="FormSubmission",
                object_id=submission.id,
                user=user_or_none,
                user_ip=get_client_ip(request),
                comments="Form submitted" + (" (anonymous)" if is_anonymous else ""),
            )

            create_approval_tasks(submission)

            # ── Success page / redirect routing ────────────────────────────
            form_data = submission.form_data or {}

            # 1. Check conditional redirect rules first
            if form_def.success_redirect_rules:
                from .conditions import evaluate_conditions

                rules = form_def.success_redirect_rules
                if isinstance(rules, list):
                    for rule in rules:
                        url = rule.pop("url", "")
                        if url and evaluate_conditions(rule, form_data):
                            rule["url"] = url  # restore for future evals
                            return redirect(_pipe_answer_tokens(url, form_data))
                        rule["url"] = url

            # 2. Static redirect URL
            if form_def.success_redirect_url:
                return redirect(
                    _pipe_answer_tokens(form_def.success_redirect_url, form_data)
                )

            # 3. Custom success message (rendered on a dedicated page)
            if form_def.success_message:
                return redirect(
                    reverse(
                        "forms_workflows:submission_success",
                        kwargs={"submission_id": submission.id},
                    )
                )

            # 4. Default behaviour
            if is_anonymous:
                return redirect(
                    "forms_workflows:public_submission_confirmation",
                )

            messages.success(request, "Form submitted successfully.")
            return redirect("forms_workflows:my_submissions")
    else:
        initial_data = draft.form_data if draft else None
        form = DynamicForm(
            form_definition=form_def, user=user_or_none, initial_data=initial_data
        )

    form_enhancements_config = json.dumps(form.get_enhancements_config())

    captcha_site_key = ""
    if form_def.enable_captcha:
        from django.conf import settings as django_settings

        captcha_site_key = getattr(
            django_settings, "FORMS_WORKFLOWS_CAPTCHA_SITE_KEY", ""
        )

    return render(
        request,
        "django_forms_workflows/form_submit.html",
        {
            "form_def": form_def,
            "form": form,
            "is_draft": draft is not None,
            "draft_id": draft.id if draft else None,
            "form_enhancements_config": form_enhancements_config,
            "is_anonymous": is_anonymous,
            "captcha_site_key": captcha_site_key,
        },
    )


@xframe_options_exempt
def form_embed(request, slug):
    """Render a form in a minimal layout for iframe embedding."""
    form_def = get_object_or_404(FormDefinition, slug=slug, is_active=True)

    if not form_def.embed_enabled:
        return HttpResponseForbidden("Embedding is not enabled for this form.")

    is_anonymous = not request.user.is_authenticated

    # Authentication gate (same as form_submit but no redirect — just 403)
    if not is_anonymous:
        if not user_can_view_form(request.user, form_def) or not user_can_submit_form(
            request.user, form_def
        ):
            return HttpResponseForbidden("Permission denied.")

    # Submission controls
    if form_def.close_date and timezone.now() >= form_def.close_date:
        return render(
            request,
            "django_forms_workflows/embed_success.html",
            {
                "form_def": form_def,
                "success_message": "This form is no longer accepting submissions.",
            },
        )
    if form_def.max_submissions is not None:
        count = (
            FormSubmission.objects.filter(form_definition=form_def)
            .exclude(status="draft")
            .count()
        )
        if count >= form_def.max_submissions:
            return render(
                request,
                "django_forms_workflows/embed_success.html",
                {
                    "form_def": form_def,
                    "success_message": (
                        "This form has reached its maximum number of submissions."
                    ),
                },
            )

    user_or_none = request.user if not is_anonymous else None
    theme = request.GET.get("theme", "")
    # Sanitise accent_color to a valid hex colour to prevent CSS injection
    _raw_accent = request.GET.get("accent_color", "")
    accent_color = (
        _raw_accent if re.fullmatch(r"#[0-9a-fA-F]{3,8}", _raw_accent) else ""
    )

    if request.method == "POST":
        # Rate limit anonymous
        if is_anonymous and not check_rate_limit(request, slug):
            return render(
                request,
                "django_forms_workflows/embed_success.html",
                {
                    "form_def": form_def,
                    "success_message": (
                        "Too many submissions. Please try again later."
                    ),
                },
            )

        stashed_files = (
            _stash_uploaded_files(request.FILES, form_def) if request.FILES else {}
        )

        form = DynamicForm(
            form_definition=form_def,
            user=user_or_none,
            data=request.POST,
            files=request.FILES,
            stashed_files=stashed_files or None,
        )

        if form.is_valid():
            submission = FormSubmission(
                form_definition=form_def,
                submitter=user_or_none,
                submission_ip=get_client_ip(request),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )
            submission.form_data = {}
            submission.save()

            submission.form_data = serialize_form_data(
                form.cleaned_data,
                submission_id=submission.pk,
                existing_file_data=stashed_files or None,
            )
            submission.form_data = _re_evaluate_calculated_fields(
                submission.form_data, form_def
            )

            # Payment gate
            if form_def.payment_enabled and form_def.payment_provider:
                submission.status = "pending_payment"
                submission.save()
                return redirect(
                    reverse(
                        "forms_workflows:payment_initiate",
                        kwargs={"submission_id": submission.id},
                    )
                )

            submission.status = "submitted"
            submission.submitted_at = timezone.now()
            submission.save()

            AuditLog.objects.create(
                action="submit",
                object_type="FormSubmission",
                object_id=submission.id,
                user=user_or_none,
                user_ip=get_client_ip(request),
                comments="Form submitted (embed)",
            )
            create_approval_tasks(submission)

            # Render inline success (no redirect)
            success_message = ""
            if form_def.success_message:
                success_message = _pipe_answer_tokens(
                    form_def.success_message, submission.form_data or {}
                )

            return render(
                request,
                "django_forms_workflows/embed_success.html",
                {
                    "form_def": form_def,
                    "submission": submission,
                    "success_message": success_message,
                    "theme": theme,
                    "accent_color": accent_color,
                },
            )
    else:
        form = DynamicForm(form_definition=form_def, user=user_or_none)

    form_enhancements_config = json.dumps(form.get_enhancements_config())

    captcha_site_key = ""
    if form_def.enable_captcha:
        from django.conf import settings as django_settings

        captcha_site_key = getattr(
            django_settings, "FORMS_WORKFLOWS_CAPTCHA_SITE_KEY", ""
        )

    response = render(
        request,
        "django_forms_workflows/form_embed.html",
        {
            "form_def": form_def,
            "form": form,
            "form_enhancements_config": form_enhancements_config,
            "captcha_site_key": captcha_site_key,
            "theme": theme,
            "accent_color": accent_color,
        },
    )

    # Handle cross-origin CSRF cookie for third-party iframe
    if hasattr(request, "META"):
        from django.conf import settings as django_settings

        response.set_cookie(
            django_settings.CSRF_COOKIE_NAME,
            request.META.get("CSRF_COOKIE", ""),
            samesite="None",
            secure=True,
            httponly=False,
        )

    return response


def _pipe_answer_tokens(text, form_data):
    """Replace {field_name} tokens in *text* with values from *form_data*.

    Unresolved tokens are replaced with an empty string so the output is
    always safe to display or use as a URL.
    """
    import re

    def _repl(m):
        val = form_data.get(m.group(1), "")
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        return str(val)

    return re.sub(r"\{(\w+)\}", _repl, text)


def submission_success(request, submission_id):
    """Custom success page with answer-piped content."""
    submission = get_object_or_404(FormSubmission, id=submission_id)
    form_def = submission.form_definition
    form_data = submission.form_data or {}

    rendered_message = _pipe_answer_tokens(form_def.success_message, form_data)

    return render(
        request,
        "django_forms_workflows/submission_success.html",
        {
            "form_def": form_def,
            "submission": submission,
            "success_message": rendered_message,
        },
    )


def public_submission_confirmation(request):
    """Thank-you page shown after an anonymous (public) form submission."""
    return render(request, "django_forms_workflows/public_submission_confirmation.html")


@require_http_methods(["POST"])
def form_auto_save(request, slug):
    """Auto-save form draft via AJAX (authenticated users only)."""
    if not request.user.is_authenticated:
        return JsonResponse(
            {"success": False, "error": "Login required for draft saving"},
            status=403,
        )
    form_def = get_object_or_404(FormDefinition, slug=slug, is_active=True)

    # Check permissions — view access is a prerequisite for submit access
    if not user_can_view_form(request.user, form_def) or not user_can_submit_form(
        request.user, form_def
    ):
        return JsonResponse(
            {"success": False, "error": "Permission denied"}, status=403
        )

    try:
        # Parse JSON data and strip any browser/Django meta-keys that should
        # never be stored as form data (defense-in-depth alongside the JS fix).
        _auto_save_skip = {"csrfmiddlewaretoken", "save_draft", "submit"}
        raw = json.loads(request.body)
        data = {k: v for k, v in raw.items() if k not in _auto_save_skip}

        # Upsert draft.  We avoid update_or_create so we can set
        # _skip_change_history *before* save() fires the signal.
        try:
            draft = FormSubmission.objects.get(
                form_definition=form_def,
                submitter=request.user,
                status="draft",
            )
        except FormSubmission.DoesNotExist:
            draft = FormSubmission(
                form_definition=form_def,
                submitter=request.user,
                status="draft",
            )

        draft.form_data = data
        draft.submission_ip = get_client_ip(request)
        draft.user_agent = request.META.get("HTTP_USER_AGENT", "")

        # Skip audit logging for auto-saves — they fire every
        # ~30 seconds and would generate excessive AuditLog /
        # ChangeHistory rows.  The draft is already persisted;
        # a proper AuditLog entry is written on explicit "Save Draft"
        # or "Submit".
        draft._skip_change_history = True
        draft.save()

        return JsonResponse(
            {
                "success": True,
                "message": "Draft saved",
                "draft_id": draft.id,
                "saved_at": draft.created_at.isoformat(),
            }
        )

    except Exception:
        logger.error(f"Auto-save error for form {slug}", exc_info=True)
        return JsonResponse(
            {"success": False, "error": "An internal error occurred."}, status=500
        )


@login_required
def my_submissions(request):
    """View user's submissions, optionally filtered by form category.

    Accepts an optional ``?category=<slug>`` query parameter to narrow the
    submission list to a single form category.  The template also receives a
    ``category_counts`` list so the UI can render a filter-pill bar showing
    how many submissions belong to each category.
    """
    # Include submissions the user owns OR submissions for forms where the user
    # is in reviewer_groups OR admin_groups (both mean "can view all submissions").
    user_groups = request.user.groups.all()
    privileged_form_ids = (
        FormDefinition.objects.filter(
            models.Q(reviewer_groups__in=user_groups)
            | models.Q(admin_groups__in=user_groups)
        )
        .values_list("id", flat=True)
        .distinct()
    )
    base_submissions = FormSubmission.objects.filter(
        models.Q(submitter=request.user)
        | models.Q(
            form_definition__in=privileged_form_ids,
            status__in=[
                "submitted",
                "pending_approval",
                "approved",
                "rejected",
                "withdrawn",
            ],
        )
    ).distinct()

    # --- Category counts (for the filter bar) ---
    raw_counts = (
        base_submissions.values(
            "form_definition__category__pk",
            "form_definition__category__name",
            "form_definition__category__slug",
            "form_definition__category__icon",
            "form_definition__category__order",
        )
        .annotate(count=models.Count("id"))
        .order_by(
            "form_definition__category__order",
            "form_definition__category__name",
        )
    )

    category_counts = []
    for row in raw_counts:
        slug = row["form_definition__category__slug"]
        if slug:  # only include categorised entries in the filter bar
            category_counts.append(
                {
                    "name": row["form_definition__category__name"],
                    "slug": slug,
                    "icon": row["form_definition__category__icon"] or "",
                    "count": row["count"],
                }
            )

    total_submissions_count = base_submissions.count()

    # --- Apply optional category filter ---
    category_slug = request.GET.get("category", "").strip()
    active_category = None

    if category_slug:
        submissions = base_submissions.filter(
            form_definition__category__slug=category_slug
        )
        active_category = next(
            (c for c in category_counts if c["slug"] == category_slug), None
        )
    else:
        submissions = base_submissions

    # --- Form counts within the active category (for the form-level filter bar) ---
    form_slug = request.GET.get("form", "").strip()
    form_counts = []
    active_form = None

    if category_slug:
        raw_form_counts = (
            submissions.values(
                "form_definition__pk",
                "form_definition__name",
                "form_definition__slug",
            )
            .annotate(count=models.Count("id"))
            .order_by("form_definition__name")
        )
        form_counts = [
            {
                "name": r["form_definition__name"],
                "slug": r["form_definition__slug"],
                "count": r["count"],
            }
            for r in raw_form_counts
            if r["form_definition__slug"]
        ]
        if form_slug:
            submissions = submissions.filter(form_definition__slug=form_slug)
            active_form = next((f for f in form_counts if f["slug"] == form_slug), None)

    # Check if any submissions support bulk export (fast EXISTS)
    any_exportable = base_submissions.filter(
        form_definition__workflows__allow_bulk_export=True
    ).exists()

    # Extra form-field columns (only when a specific form is selected)
    form_fields = []
    if form_slug:
        try:
            fd = FormDefinition.objects.get(slug=form_slug)
            form_fields = list(
                FormField.objects.filter(form_definition=fd)
                .exclude(field_type__in=["section", "file", "multifile", "signature"])
                .order_by("order")
                .values("field_name", "field_label")
            )
        except FormDefinition.DoesNotExist:
            logger.debug(
                "FormDefinition not found when loading export fields for submission list"
            )

    # Compute default sort column index for DataTables (submitted_at)
    _exp_off = 1 if any_exportable else 0
    _cat_off = 0 if category_slug else 1
    # columns: [checkbox?] id actions [category?] form status submitted_at [extra…]
    default_sort_col = _exp_off + 2 + _cat_off + 2  # index of submitted_at

    return render(
        request,
        "django_forms_workflows/my_submissions.html",
        {
            "category_counts": category_counts,
            "active_category": active_category,
            "category_slug": category_slug,
            "total_submissions_count": total_submissions_count,
            "form_counts": form_counts,
            "form_slug": form_slug,
            "active_form": active_form,
            "any_exportable": any_exportable,
            "form_fields": form_fields,
            "default_sort_col": default_sort_col,
        },
    )


@login_required
def submission_detail(request, submission_id):
    """View submission details"""
    submission = get_object_or_404(FormSubmission, id=submission_id)

    # Check permissions - user must be submitter, approver, admin, or reviewer
    form_def_early = submission.form_definition
    is_reviewer = request.user.groups.filter(
        id__in=form_def_early.reviewer_groups.all()
    ).exists()
    can_view = (
        submission.submitter == request.user
        or request.user.is_superuser
        or user_can_approve(request.user, submission)
        or request.user.groups.filter(id__in=form_def_early.admin_groups.all()).exists()
        or is_reviewer
    )

    if not can_view:
        return HttpResponseForbidden(
            "You don't have permission to view this submission."
        )

    # Get approval tasks with related objects for efficient rendering
    approval_tasks = submission.approval_tasks.select_related(
        "workflow_stage", "assigned_to", "assigned_group", "completed_by"
    ).order_by("stage_number", "-created_at")

    # Build stage groups if any tasks are linked to a workflow stage.
    # Separate parent-workflow tasks from sub-workflow tasks so they
    # render in distinct sections.
    stage_groups = None
    sub_workflow_groups = None
    workflow = getattr(submission.form_definition, "workflow", None)
    if approval_tasks.filter(workflow_stage__isnull=False).exists():
        # Determine which stage IDs belong to the parent workflow
        parent_stage_ids = set()
        if workflow:
            parent_stage_ids = set(workflow.stages.values_list("id", flat=True))

        parent_groups: dict = {}
        sub_groups: dict = {}
        collapse = bool(
            workflow and getattr(workflow, "collapse_parallel_stages", False)
        )
        for task in approval_tasks:
            if not task.workflow_stage_id:
                continue
            is_parent = task.workflow_stage_id in parent_stage_ids
            target = parent_groups if is_parent else sub_groups
            # When collapse_parallel_stages is enabled on the workflow,
            # parent tasks at the same order number share a single card.
            # Otherwise (default) every distinct workflow_stage gets its own card.
            if is_parent and collapse:
                stage_key = task.stage_number or 0
            else:
                stage_key = task.workflow_stage_id
            if stage_key not in target:
                stage_name = (
                    task.workflow_stage.name
                    if task.workflow_stage
                    else f"Stage {task.stage_number or 0}"
                )
                target[stage_key] = {
                    "number": task.stage_number or 0,
                    "name": stage_name,
                    "approval_logic": (
                        task.workflow_stage.approval_logic
                        if task.workflow_stage
                        else "all"
                    ),
                    "tasks": [],
                    "approved_count": 0,
                    "rejected_count": 0,
                    "total_count": 0,
                    "has_multiple_stages": False,
                }
            group = target[stage_key]
            # Enrich task with its stage name for the "Step" column
            task.step_display_name = (
                task.workflow_stage.name
                if task.workflow_stage
                else f"Stage {task.stage_number or 0}"
            )
            group["tasks"].append(task)
            group["total_count"] += 1
            if task.status == "approved":
                group["approved_count"] += 1
            elif task.status == "rejected":
                group["rejected_count"] += 1

        # When collapsed, detect groups that merged multiple parallel stages
        # so the template can omit the per-stage name in the card header.
        if collapse:
            for group in parent_groups.values():
                stage_ids = {t.workflow_stage_id for t in group["tasks"]}
                if len(stage_ids) > 1:
                    group["has_multiple_stages"] = True

        if parent_groups:
            stage_groups = sorted(parent_groups.values(), key=lambda x: x["number"])
        if sub_groups:
            sub_workflow_groups = sorted(sub_groups.values(), key=lambda x: x["number"])
            # Resolve the user-facing section label from SubWorkflowDefinition
            if workflow:
                sub_wf_config = getattr(workflow, "sub_workflow_config", None)
                if sub_wf_config and sub_wf_config.section_label:
                    for sg in sub_workflow_groups:
                        sg["section_label"] = sub_wf_config.section_label
                elif sub_wf_config:
                    sg_name = sub_wf_config.sub_workflow.form_definition.name
                    for sg in sub_workflow_groups:
                        sg["section_label"] = sg_name

    # Build per-instance stage groups so each sub-workflow instance is
    # rendered as its own section with stage cards (matching parent style).
    sub_workflow_instance_stages = None
    sub_wfs = list(
        submission.sub_workflows.select_related("definition")
        .prefetch_related(
            "approval_tasks__workflow_stage",
            "approval_tasks__assigned_to",
            "approval_tasks__assigned_group",
            "approval_tasks__completed_by",
        )
        .order_by("index")
    )
    if sub_wfs:
        instance_sections = []
        for swf in sub_wfs:
            tasks = list(
                swf.approval_tasks.select_related("workflow_stage").order_by(
                    "stage_number", "id"
                )
            )
            # Group tasks by workflow_stage (like parent stage_groups)
            stage_map: dict = {}
            inst_approved = 0
            inst_rejected = 0
            inst_total = 0
            for task in tasks:
                stage_key = task.workflow_stage_id or f"flat_{task.stage_number}"
                if stage_key not in stage_map:
                    stage_name = (
                        task.workflow_stage.name
                        if task.workflow_stage
                        else f"Step {task.stage_number or 1}"
                    )
                    approval_logic = (
                        task.workflow_stage.approval_logic
                        if task.workflow_stage
                        else "all"
                    )
                    stage_map[stage_key] = {
                        "number": task.stage_number or 0,
                        "name": stage_name,
                        "approval_logic": approval_logic,
                        "tasks": [],
                        "approved_count": 0,
                        "rejected_count": 0,
                        "total_count": 0,
                    }
                stage_map[stage_key]["tasks"].append(task)
                stage_map[stage_key]["total_count"] += 1
                inst_total += 1
                if task.status == "approved":
                    stage_map[stage_key]["approved_count"] += 1
                    inst_approved += 1
                elif task.status == "rejected":
                    stage_map[stage_key]["rejected_count"] += 1
                    inst_rejected += 1

            instance_sections.append(
                {
                    "instance": swf,
                    "label": swf.label,
                    "stage_groups": sorted(
                        stage_map.values(), key=lambda x: x["number"]
                    ),
                    "approved_count": inst_approved,
                    "rejected_count": inst_rejected,
                    "total_count": inst_total,
                }
            )
        sub_workflow_instance_stages = instance_sections

    # Resolve fresh presigned URLs for any file-upload fields
    form_data = _resolve_form_data_urls(submission.form_data)
    form_data_ordered = _build_ordered_form_data(submission, form_data)

    # Resolve presigned URLs for model-level attachments (e.g. migrated SP files)
    resolved_attachments = _resolve_attachments(submission.attachments)

    # Collect stage-scoped field names so the template can exclude them from
    # the main "Form Data" table and render them in their own sections.
    approval_field_names = list(
        submission.form_definition.fields.filter(
            workflow_stage__isnull=False
        ).values_list("field_name", flat=True)
    )
    approval_step_sections = _build_approval_step_sections(submission)

    # Elevated viewers (approvers / admins / superusers) may download PDFs of
    # post_approval forms even before the submission is fully approved.
    form_def = submission.form_definition
    can_approve_pdf = (
        request.user.is_superuser
        or user_can_approve(request.user, submission)
        or request.user.groups.filter(id__in=form_def.admin_groups.all()).exists()
        or is_reviewer
    )

    # Privacy: hide approval history from the submitter when configured.
    # Approvers, admins, and reviewers always see the full history.
    is_submitter_only = (
        submission.submitter_id is not None
        and submission.submitter_id == request.user.pk
        and not request.user.is_superuser
        and not user_can_approve(request.user, submission)
        and not request.user.groups.filter(id__in=form_def.admin_groups.all()).exists()
        and not is_reviewer
    )
    hide_approval_history = bool(
        workflow and workflow.hide_approval_history and is_submitter_only
    )

    return render(
        request,
        "django_forms_workflows/submission_detail.html",
        {
            "submission": submission,
            "approval_tasks": approval_tasks,
            "stage_groups": stage_groups,
            "workflow_name_label": (
                workflow.name_label if workflow and workflow.name_label else None
            ),
            "sub_workflow_groups": sub_workflow_groups,
            "sub_workflow_instance_stages": sub_workflow_instance_stages,
            "form_data": form_data,
            "form_data_ordered": form_data_ordered,
            "resolved_attachments": resolved_attachments,
            "approval_field_names": approval_field_names,
            "approval_step_sections": approval_step_sections,
            "can_approve_pdf": can_approve_pdf,
            "hide_approval_history": hide_approval_history,
        },
    )


@login_required
def approval_inbox(request):
    """View pending approvals, optionally filtered by form category.

    Accepts an optional ``?category=<slug>`` query parameter to narrow the
    task list to a single form category.  The template also receives a
    ``category_counts`` list so the UI can render a filter-pill bar showing
    how many pending tasks belong to each category.
    """
    # Build base queryset of accessible pending tasks (no select_related yet
    # so the values()/annotate() path stays lightweight).
    if request.user.is_superuser:
        base_tasks = ApprovalTask.objects.filter(status="pending")
    else:
        user_groups = request.user.groups.all()
        base_tasks = ApprovalTask.objects.filter(status="pending").filter(
            models.Q(assigned_to=request.user)
            | models.Q(assigned_group__in=user_groups)
        )

    # --- Category counts (for the filter bar) ---
    raw_counts = (
        base_tasks.values(
            "submission__form_definition__category__pk",
            "submission__form_definition__category__name",
            "submission__form_definition__category__slug",
            "submission__form_definition__category__icon",
            "submission__form_definition__category__order",
        )
        .annotate(count=models.Count("id"))
        .order_by(
            "submission__form_definition__category__order",
            "submission__form_definition__category__name",
        )
    )

    category_counts = []
    for row in raw_counts:
        slug = row["submission__form_definition__category__slug"]
        if slug:  # only include categorised entries in the filter bar
            category_counts.append(
                {
                    "name": row["submission__form_definition__category__name"],
                    "slug": slug,
                    "icon": row["submission__form_definition__category__icon"] or "",
                    "count": row["count"],
                }
            )

    total_tasks_count = base_tasks.count()

    # --- Completed count for cross-tab badge (must mirror completed_approvals) ---
    _history_statuses = ["approved", "pending_approval", "rejected", "withdrawn"]
    if request.user.is_superuser:
        completed_count = FormSubmission.objects.filter(
            status__in=_history_statuses
        ).count()
    else:
        _completed_task_sub_ids = (
            ApprovalTask.objects.filter(
                models.Q(assigned_to=request.user)
                | models.Q(assigned_group__in=user_groups),
                status__in=["pending", "approved", "rejected"],
            )
            .values_list("submission_id", flat=True)
            .distinct()
        )
        _privileged_form_ids = (
            FormDefinition.objects.filter(
                models.Q(reviewer_groups__in=user_groups)
                | models.Q(admin_groups__in=user_groups)
            )
            .values_list("id", flat=True)
            .distinct()
        )
        completed_count = (
            FormSubmission.objects.filter(
                models.Q(id__in=_completed_task_sub_ids)
                | models.Q(
                    form_definition__in=_privileged_form_ids,
                    status__in=_history_statuses,
                )
            )
            .distinct()
            .count()
        )

    # --- Apply optional category filter ---
    category_slug = request.GET.get("category", "").strip()
    active_category = None

    if category_slug:
        display_tasks = base_tasks.filter(
            submission__form_definition__category__slug=category_slug
        )
        active_category = next(
            (c for c in category_counts if c["slug"] == category_slug), None
        )
    else:
        display_tasks = base_tasks

    # --- Form counts within the active category (for the form-level filter bar) ---
    form_slug = request.GET.get("form", "").strip()
    form_counts = []
    active_form = None

    if category_slug:
        raw_form_counts = (
            display_tasks.values(
                "submission__form_definition__pk",
                "submission__form_definition__name",
                "submission__form_definition__slug",
            )
            .annotate(count=models.Count("id"))
            .order_by("submission__form_definition__name")
        )
        form_counts = [
            {
                "name": r["submission__form_definition__name"],
                "slug": r["submission__form_definition__slug"],
                "count": r["count"],
            }
            for r in raw_form_counts
            if r["submission__form_definition__slug"]
        ]
        if form_slug:
            display_tasks = display_tasks.filter(
                submission__form_definition__slug=form_slug
            )
            active_form = next((f for f in form_counts if f["slug"] == form_slug), None)

    # --- Form fields for the column picker (only when a specific form is active) ---
    form_fields = []
    if form_slug and active_form:
        try:
            form_def_obj = FormDefinition.objects.get(slug=form_slug)
            form_fields = list(
                FormField.objects.filter(form_definition=form_def_obj)
                .exclude(field_type__in=["section", "file", "multifile", "signature"])
                .order_by("order")
                .values("field_name", "field_label")
            )
        except FormDefinition.DoesNotExist:
            logger.debug(
                "FormDefinition not found when loading export fields for approval inbox"
            )

    any_exportable = base_tasks.filter(
        submission__form_definition__workflows__allow_bulk_export=True
    ).exists()
    any_pdf_exportable = base_tasks.filter(
        submission__form_definition__workflows__allow_bulk_pdf_export=True
    ).exists()

    # Compute default sort column index for DataTables (submitted_at)
    _exp_off = 1 if (any_exportable or any_pdf_exportable) else 0
    _cat_off = 0 if category_slug else 1
    # columns: [checkbox?] [category?] actions form submitter stage step_num assigned submitted_at
    default_sort_col = _exp_off + _cat_off + 1 + 5  # index of submitted_at

    return render(
        request,
        "django_forms_workflows/approval_inbox.html",
        {
            "category_counts": category_counts,
            "active_category": active_category,
            "category_slug": category_slug,
            "total_tasks_count": total_tasks_count,
            "completed_count": completed_count,
            "form_counts": form_counts,
            "form_slug": form_slug,
            "active_form": active_form,
            "form_fields": form_fields,
            "any_exportable": any_exportable,
            "any_pdf_exportable": any_pdf_exportable,
            "default_sort_col": default_sort_col,
        },
    )


@login_required
def approve_submission(request, task_id):
    """Approve or reject a submission"""
    task = get_object_or_404(ApprovalTask, id=task_id)
    submission = task.submission
    form_def = submission.form_definition

    # For sub-workflow tasks, approval step fields belong to the sub-workflow's
    # form definition, not the parent form.  Use field_form_def for all field
    # lookups and ApprovalStepForm construction; keep form_def for everything else.
    if task.sub_workflow_instance_id:
        field_form_def = (
            task.sub_workflow_instance.definition.sub_workflow.form_definition
        )
    else:
        field_form_def = form_def

    # Check permission
    can_approve = (
        task.assigned_to == request.user
        or (task.assigned_group and task.assigned_group in request.user.groups.all())
        or request.user.is_superuser
    )

    if not can_approve:
        messages.error(request, "You don't have permission to approve this.")
        return redirect("forms_workflows:approval_inbox")

    if task.status != "pending":
        messages.warning(request, "This task has already been processed.")
        return redirect("forms_workflows:approval_inbox")

    # Check if this form has stage-scoped fields for the current task.
    has_approval_step_fields = (
        task.workflow_stage_id
        and field_form_def.fields.filter(workflow_stage_id=task.workflow_stage_id)
        .exclude(field_type="section")
        .exists()
    )

    # Check if this stage allows editing the original submission data.
    allow_edit_form_data = bool(
        task.workflow_stage_id
        and task.workflow_stage
        and task.workflow_stage.allow_edit_form_data
    )

    approval_step_form = None
    editable_form = None

    if request.method == "POST":
        decision = request.POST.get("decision")
        comments = request.POST.get("comments", "")

        if decision not in ["approve", "reject", "send_back"]:
            messages.error(request, "Invalid decision.")
            return redirect("forms_workflows:approve_submission", task_id=task_id)

        # Send-back path — no approval-step-field validation needed.
        if decision == "send_back":
            send_back_reason = request.POST.get("send_back_reason", "").strip()
            send_back_stage_id = request.POST.get("send_back_stage_id", "").strip()
            if not send_back_reason:
                messages.error(
                    request, "A reason is required when sending back for correction."
                )
                return redirect("forms_workflows:approve_submission", task_id=task_id)
            if not send_back_stage_id:
                messages.error(request, "Please select a stage to send back to.")
                return redirect("forms_workflows:approve_submission", task_id=task_id)

            from .models import WorkflowStage
            from .workflow_engine import handle_send_back, handle_sub_workflow_send_back

            try:
                target_stage = WorkflowStage.objects.get(pk=send_back_stage_id)
            except WorkflowStage.DoesNotExist:
                messages.error(request, "Invalid target stage selected.")
                return redirect("forms_workflows:approve_submission", task_id=task_id)

            # Verify the target stage belongs to the same workflow and is truly prior.
            current_stage = task.workflow_stage
            if (
                current_stage is None
                or target_stage.workflow_id != current_stage.workflow_id
            ):
                messages.error(
                    request, "Target stage does not belong to this workflow."
                )
                return redirect("forms_workflows:approve_submission", task_id=task_id)
            if target_stage.order >= current_stage.order:
                messages.error(request, "You can only send back to a prior stage.")
                return redirect("forms_workflows:approve_submission", task_id=task_id)

            task.status = "returned"
            task.decision = "send_back"
            task.comments = send_back_reason
            task.completed_by = request.user
            task.completed_at = timezone.now()
            task.save()

            if task.sub_workflow_instance_id:
                handle_sub_workflow_send_back(task, target_stage)
            else:
                handle_send_back(submission, task, target_stage)

            AuditLog.objects.create(
                action="send_back",
                object_type="FormSubmission",
                object_id=submission.id,
                user=request.user,
                user_ip=get_client_ip(request),
                changes={
                    "task_id": task.id,
                    "target_stage": target_stage.name,
                    "reason": send_back_reason,
                },
            )
            messages.success(
                request,
                f'Submission returned to "{target_stage.name}" for correction.',
            )
            return redirect("forms_workflows:approval_inbox")

        # If there are approval step fields and decision is approve, validate them
        if has_approval_step_fields and decision == "approve":
            from .forms import ApprovalStepForm

            approval_step_form = ApprovalStepForm(
                form_definition=field_form_def,
                submission=submission,
                approval_task=task,
                user=request.user,
                data=request.POST,
                files=request.FILES,
            )

            if not approval_step_form.is_valid():
                messages.error(
                    request, "Please correct the errors in the approval fields."
                )
                _fd = _resolve_form_data_urls(submission.form_data)
                _ctx = {
                    "task": task,
                    "submission": submission,
                    "approval_step_form": approval_step_form,
                    "has_approval_step_fields": has_approval_step_fields,
                    "allow_edit_form_data": allow_edit_form_data,
                    "form_data": _fd,
                    "form_data_ordered": _build_ordered_form_data(submission, _fd),
                    "resolved_attachments": _resolve_attachments(
                        submission.attachments
                    ),
                }
                if allow_edit_form_data:
                    from .forms import DynamicForm

                    _ctx["editable_form"] = DynamicForm(
                        form_def,
                        user=request.user,
                        initial_data=submission.form_data,
                        data=request.POST,
                        files=request.FILES,
                    )
                return render(
                    request,
                    "django_forms_workflows/approve.html",
                    _ctx,
                )

            # Update submission form_data with approval step fields.
            # For sub-workflow tasks, remap generic field names to indexed names
            # (e.g. payment_type → payment_type_1) so multiple payments don't overwrite each other.
            updated_data = approval_step_form.get_updated_form_data()
            if task.sub_workflow_instance_id:
                idx = task.sub_workflow_instance.index
                stage_field_names = set(
                    field_form_def.fields.filter(
                        workflow_stage_id=task.workflow_stage_id
                    ).values_list("field_name", flat=True)
                )
                updated_data = {
                    (f"{k}_{idx}" if k in stage_field_names else k): v
                    for k, v in updated_data.items()
                }
            submission.form_data = updated_data
            submission.save()

        # If this stage allows editing form data, validate and merge on approve
        if allow_edit_form_data and decision == "approve":
            from .forms import DynamicForm

            editable_form = DynamicForm(
                form_def,
                user=request.user,
                initial_data=submission.form_data,
                data=request.POST,
                files=request.FILES,
            )
            if not editable_form.is_valid():
                messages.error(
                    request, "Please correct the errors in the submission data."
                )
                _fd = _resolve_form_data_urls(submission.form_data)
                _ctx = {
                    "task": task,
                    "submission": submission,
                    "editable_form": editable_form,
                    "allow_edit_form_data": allow_edit_form_data,
                    "has_approval_step_fields": has_approval_step_fields,
                    "form_data": _fd,
                    "form_data_ordered": _build_ordered_form_data(submission, _fd),
                    "resolved_attachments": _resolve_attachments(
                        submission.attachments
                    ),
                }
                if has_approval_step_fields:
                    from .forms import ApprovalStepForm

                    _ctx["approval_step_form"] = ApprovalStepForm(
                        form_definition=field_form_def,
                        submission=submission,
                        approval_task=task,
                        user=request.user,
                        data=request.POST,
                        files=request.FILES,
                    )
                return render(
                    request,
                    "django_forms_workflows/approve.html",
                    _ctx,
                )
            # Merge edited fields into submission data, preserving any
            # approval-step fields that are not part of the original form.
            edited_data = serialize_form_data(
                editable_form.cleaned_data, submission_id=submission.id
            )
            # Log the form_data diff before overwriting
            old_form_data = dict(submission.form_data)
            submission.form_data.update(edited_data)
            from .models import ChangeHistory

            ChangeHistory.log_json_diff(
                submission,
                "form_data",
                old_form_data,
                submission.form_data,
                user=request.user,
            )
            submission.save()

        # Update task
        task.status = "approved" if decision == "approve" else "rejected"
        task.completed_by = request.user
        task.completed_at = timezone.now()
        task.comments = comments
        task.decision = decision
        task.save()

        # Update submission status
        workflow = form_def.workflow

        if task.sub_workflow_instance_id:
            # Sub-workflow task — dispatch to sub-workflow engine
            from .workflow_engine import (
                handle_sub_workflow_approval,
                handle_sub_workflow_rejection,
            )

            if decision == "reject":
                handle_sub_workflow_rejection(task)
            else:
                handle_sub_workflow_approval(task)
        elif decision == "reject":
            from .workflow_engine import handle_rejection

            handle_rejection(submission, task, workflow)
        else:
            # Approval - check workflow logic
            from .workflow_engine import handle_approval

            handle_approval(submission, task, workflow)

        # Log audit
        AuditLog.objects.create(
            action="approve" if decision == "approve" else "reject",
            object_type="FormSubmission",
            object_id=submission.id,
            user=request.user,
            user_ip=get_client_ip(request),
            changes={"task_id": task.id, "comments": comments},
        )

        action_label = "approved" if decision == "approve" else "rejected"
        messages.success(request, f"Submission {action_label} successfully.")
        return redirect("forms_workflows:approval_inbox")

    # GET request - create the approval step form if needed
    if has_approval_step_fields:
        from .forms import ApprovalStepForm

        approval_step_form = ApprovalStepForm(
            form_definition=field_form_def,
            submission=submission,
            approval_task=task,
            user=request.user,
        )

    # Build the editable form data form if the stage allows it
    if allow_edit_form_data:
        from .forms import DynamicForm

        editable_form = DynamicForm(
            form_def,
            user=request.user,
            initial_data=submission.form_data,
        )

    # Build approval progress context from workflow stages
    all_approval_field_names: list = []
    approval_stages: list = []
    parallel_tasks: list = []
    workflow_mode = "none"
    workflow = form_def.workflow

    if workflow:
        stages = list(workflow.stages.order_by("order"))

        if stages:
            # -------- staged workflow --------
            workflow_mode = "staged"
            current_stage_number = task.stage_number or 1
            all_sub_tasks = list(
                submission.approval_tasks.select_related(
                    "workflow_stage", "assigned_group", "assigned_to"
                ).all()
            )
            tasks_by_stage: dict = {}
            for t in all_sub_tasks:
                key = t.workflow_stage_id
                tasks_by_stage.setdefault(key, []).append(t)

            for i, stage in enumerate(stages, start=1):
                stage_tasks = tasks_by_stage.get(stage.id, [])
                approved_count = sum(1 for t in stage_tasks if t.status == "approved")
                approval_stages.append(
                    {
                        "number": i,
                        "name": stage.name,
                        "approval_logic": stage.approval_logic,
                        "total_stages": len(stages),
                        "is_current": i == current_stage_number,
                        "is_completed": approved_count > 0
                        and not any(t.status == "pending" for t in stage_tasks),
                        "is_rejected": any(t.status == "rejected" for t in stage_tasks),
                        "is_pending": i > current_stage_number,
                        "tasks": stage_tasks,
                        "approved_count": approved_count,
                        "total_count": len(stage_tasks),
                    }
                )

            # Collect stage-scoped field names so the template can exclude
            # them from the read-only "Submission Data" table.
            all_approval_field_names = list(
                form_def.fields.filter(workflow_stage__isnull=False).values_list(
                    "field_name", flat=True
                )
            )

        elif submission.approval_tasks.filter(assigned_group__isnull=False).exists():
            # -------- legacy unstaged parallel tasks --------
            workflow_mode = "all"
            for t in submission.approval_tasks.filter(
                assigned_group__isnull=False
            ).select_related("assigned_group"):
                parallel_tasks.append(
                    {
                        "task": t,
                        "group_name": t.assigned_group.name
                        if t.assigned_group
                        else "—",
                        "is_current": t.id == task.id,
                    }
                )

    # Resolve fresh presigned URLs for any file-upload fields
    form_data = _resolve_form_data_urls(submission.form_data)
    form_data_ordered = _build_ordered_form_data(submission, form_data)

    # Resolve presigned URLs for model-level attachments (e.g. migrated SP files)
    resolved_attachments = _resolve_attachments(submission.attachments)

    # Resolve the approve-button label from the stage (each parallel stage
    # has its own label, so no per-group override is needed).
    approve_label = (
        task.workflow_stage.approve_label
        if task.workflow_stage and task.workflow_stage.approve_label
        else ""
    )

    # Build the list of prior stages the approver may send back to.
    # For sub-workflow tasks we look at the sub-workflow's own stage set.
    send_back_stages: list = []
    if task.workflow_stage and task.workflow_stage.allow_send_back is not None:
        # current task's stage is NOT a send-back target for itself; only
        # stages with a *lower* order value and allow_send_back=True qualify.
        current_order = task.workflow_stage.order
        if task.sub_workflow_instance_id:
            sub_wf = task.sub_workflow_instance.definition.sub_workflow
            send_back_stages = list(
                sub_wf.stages.filter(
                    allow_send_back=True, order__lt=current_order
                ).order_by("order")
            )
        else:
            wf_for_send_back = task.workflow_stage.workflow
            send_back_stages = list(
                wf_for_send_back.stages.filter(
                    allow_send_back=True, order__lt=current_order
                ).order_by("order")
            )

    # Build completed approval step sections so stage 2+ approvers can see
    # prior approval step field data (with file URLs resolved).
    approval_step_sections = _build_approval_step_sections(submission)

    return render(
        request,
        "django_forms_workflows/approve.html",
        {
            "task": task,
            "submission": submission,
            "approval_step_form": approval_step_form,
            "has_approval_step_fields": has_approval_step_fields,
            "allow_edit_form_data": allow_edit_form_data,
            "editable_form": editable_form,
            "approval_field_names": all_approval_field_names,
            # staged
            "approval_stages": approval_stages,
            "current_stage_number": task.stage_number or 1,
            "total_stages": len(approval_stages),
            # parallel legacy
            "parallel_tasks": parallel_tasks,
            # shared
            "workflow_mode": workflow_mode,
            "form_data": form_data,
            "form_data_ordered": form_data_ordered,
            "resolved_attachments": resolved_attachments,
            # prior approval step responses
            "approval_step_sections": approval_step_sections,
            # custom button label
            "approve_label": approve_label,
            # sub-workflow context (None for regular tasks)
            "sub_workflow_instance": task.sub_workflow_instance,
            # send-back targets (empty list → panel hidden in template)
            "send_back_stages": send_back_stages,
            # reassign option
            "allow_reassign": bool(
                task.workflow_stage and task.workflow_stage.allow_reassign
            ),
            # hide the public decision comment textarea
            "hide_comment_field": bool(
                task.workflow_stage and task.workflow_stage.hide_comment_field
            ),
        },
    )


@login_required
def reassign_task(request, task_id):
    """Reassign an approval task to another member of the stage's approval groups."""
    from django.contrib.auth import get_user_model

    user_model = get_user_model()
    task = get_object_or_404(ApprovalTask, id=task_id)

    if task.status != "pending":
        messages.warning(request, "This task has already been processed.")
        return redirect("forms_workflows:approval_inbox")

    stage = task.workflow_stage
    if not stage or not stage.allow_reassign:
        messages.error(request, "Reassignment is not enabled for this approval step.")
        return redirect("forms_workflows:approval_inbox")

    # Permission: current assignee, any member of the stage's reassignment
    # groups (falls back to approval groups), or superuser can reassign.
    reassign_groups = list(stage.get_reassignment_groups())
    reassign_group_ids = {g.pk for g in reassign_groups}
    can_reassign = (
        task.assigned_to == request.user
        or request.user.groups.filter(pk__in=reassign_group_ids).exists()
        or request.user.is_superuser
    )
    if not can_reassign:
        messages.error(request, "You don't have permission to reassign this task.")
        return redirect("forms_workflows:approval_inbox")

    if request.method == "POST":
        new_assignee_id = request.POST.get("new_assignee_id", "").strip()
        if not new_assignee_id:
            messages.error(request, "Please select a user to reassign to.")
            return redirect("forms_workflows:reassign_task", task_id=task_id)

        try:
            new_assignee = user_model.objects.get(pk=new_assignee_id)
        except user_model.DoesNotExist:
            messages.error(request, "Selected user not found.")
            return redirect("forms_workflows:reassign_task", task_id=task_id)

        # Validate the new assignee is in one of the stage's reassignment groups
        if not new_assignee.groups.filter(pk__in=reassign_group_ids).exists():
            messages.error(
                request,
                "The selected user is not a member of any reassignment group for this stage.",
            )
            return redirect("forms_workflows:reassign_task", task_id=task_id)

        old_assignee = task.assigned_to
        task.assigned_to = new_assignee
        task.save(update_fields=["assigned_to"])

        AuditLog.objects.create(
            action="reassign",
            object_type="ApprovalTask",
            object_id=task.id,
            user=request.user,
            user_ip=get_client_ip(request),
            changes={
                "old_assignee": old_assignee.username if old_assignee else None,
                "new_assignee": new_assignee.username,
                "submission_id": task.submission_id,
            },
        )

        # Notify the new assignee
        from .workflow_engine import _notify_task_request

        _notify_task_request(task)

        messages.success(
            request,
            f"Task reassigned to {new_assignee.get_full_name() or new_assignee.username}.",
        )
        return redirect("forms_workflows:approval_inbox")

    # GET — show reassignment form with eligible users
    eligible_users = (
        user_model.objects.filter(groups__pk__in=reassign_group_ids, is_active=True)
        .distinct()
        .order_by("last_name", "first_name", "username")
    )

    return render(
        request,
        "django_forms_workflows/reassign_task.html",
        {
            "task": task,
            "submission": task.submission,
            "stage": stage,
            "eligible_users": eligible_users,
        },
    )


@login_required
def sub_workflow_detail(request, instance_id):
    """View details of a single sub-workflow instance (e.g. Payment 1)."""
    from .models import SubWorkflowInstance

    instance = get_object_or_404(SubWorkflowInstance, id=instance_id)
    submission = instance.parent_submission

    can_view = (
        submission.submitter == request.user
        or request.user.is_superuser
        or user_can_approve(request.user, submission)
        or request.user.groups.filter(
            id__in=submission.form_definition.admin_groups.all()
        ).exists()
        or request.user.groups.filter(
            id__in=submission.form_definition.reviewer_groups.all()
        ).exists()
    )
    if not can_view:
        messages.error(request, "You don't have permission to view this.")
        return redirect("forms_workflows:my_submissions")

    tasks = instance.approval_tasks.select_related(
        "assigned_to", "assigned_group", "completed_by", "workflow_stage"
    ).order_by("created_at")

    return render(
        request,
        "django_forms_workflows/sub_workflow_detail.html",
        {
            "instance": instance,
            "submission": submission,
            "tasks": tasks,
            "form_data_slice": instance.form_data_slice,
        },
    )


@login_required
def completed_approvals(request):
    """View completed submissions where the user was part of the approval workflow.

    Shows submissions with status approved/rejected/withdrawn where the current
    user had an ApprovalTask assigned directly or via group membership.  This
    provides a history / audit view for approvers, which is especially important
    for business-process workflows (e.g. PCN) where data retention matters.

    Accepts an optional ``?category=<slug>`` query parameter and an optional
    ``?form=<slug>`` query parameter for narrowing the list.
    """
    _history_statuses = ["approved", "pending_approval", "rejected", "withdrawn"]
    if request.user.is_superuser:
        base_submissions = FormSubmission.objects.filter(status__in=_history_statuses)
    else:
        user_groups = request.user.groups.all()
        # Include any submission the user has an ApprovalTask for — pending,
        # approved, or rejected — so that submissions they can see on the
        # Pending tab also appear here when filtered to "pending_approval".
        completed_task_sub_ids = (
            ApprovalTask.objects.filter(
                models.Q(assigned_to=request.user)
                | models.Q(assigned_group__in=user_groups),
                status__in=["pending", "approved", "rejected"],
            )
            .values_list("submission_id", flat=True)
            .distinct()
        )
        # reviewer_groups and admin_groups members both see all non-draft
        # submissions for their forms, including zero-approval-step forms.
        privileged_form_ids = (
            FormDefinition.objects.filter(
                models.Q(reviewer_groups__in=user_groups)
                | models.Q(admin_groups__in=user_groups)
            )
            .values_list("id", flat=True)
            .distinct()
        )
        base_submissions = FormSubmission.objects.filter(
            models.Q(id__in=completed_task_sub_ids)
            | models.Q(
                form_definition__in=privileged_form_ids,
                status__in=_history_statuses,
            )
        ).distinct()

    # --- Category counts (for the filter bar) ---
    raw_counts = (
        base_submissions.values(
            "form_definition__category__pk",
            "form_definition__category__name",
            "form_definition__category__slug",
            "form_definition__category__icon",
            "form_definition__category__order",
        )
        .annotate(count=models.Count("id"))
        .order_by(
            "form_definition__category__order",
            "form_definition__category__name",
        )
    )

    category_counts = []
    for row in raw_counts:
        slug = row["form_definition__category__slug"]
        if slug:
            category_counts.append(
                {
                    "name": row["form_definition__category__name"],
                    "slug": slug,
                    "icon": row["form_definition__category__icon"] or "",
                    "count": row["count"],
                }
            )

    # Derive from the aggregation we already ran — saves a redundant COUNT(*) query
    total_count = sum(c["count"] for c in category_counts)

    # --- Pending tasks count for cross-tab badge ---
    if request.user.is_superuser:
        pending_tasks_count = ApprovalTask.objects.filter(status="pending").count()
    else:
        pending_tasks_count = (
            ApprovalTask.objects.filter(status="pending")
            .filter(
                models.Q(assigned_to=request.user)
                | models.Q(assigned_group__in=user_groups)
            )
            .count()
        )

    # --- Apply optional category filter ---
    category_slug = request.GET.get("category", "").strip()
    active_category = None
    filtered_submissions = base_submissions

    if category_slug:
        filtered_submissions = base_submissions.filter(
            form_definition__category__slug=category_slug
        )
        active_category = next(
            (c for c in category_counts if c["slug"] == category_slug), None
        )

    # --- Form counts within the active category (for the form-level filter bar) ---
    form_slug = request.GET.get("form", "").strip()
    form_counts = []
    active_form = None

    if category_slug:
        raw_form_counts = (
            filtered_submissions.values(
                "form_definition__pk",
                "form_definition__name",
                "form_definition__slug",
            )
            .annotate(count=models.Count("id"))
            .order_by("form_definition__name")
        )
        form_counts = [
            {
                "name": r["form_definition__name"],
                "slug": r["form_definition__slug"],
                "count": r["count"],
            }
            for r in raw_form_counts
            if r["form_definition__slug"]
        ]
        if form_slug:
            active_form = next((f for f in form_counts if f["slug"] == form_slug), None)

    status_filter = request.GET.get("status", "").strip()

    # --- Bulk export flags (fast EXISTS — no full queryset fetch needed) ---
    any_exportable = base_submissions.filter(
        form_definition__workflows__allow_bulk_export=True
    ).exists()
    any_pdf_exportable = base_submissions.filter(
        form_definition__workflows__allow_bulk_pdf_export=True
    ).exists()

    # --- Form fields for column picker (when a specific form is filtered) ---
    form_fields = []
    if form_slug and active_form:
        try:
            form_def_obj = FormDefinition.objects.get(slug=form_slug)
            form_fields = list(
                FormField.objects.filter(form_definition=form_def_obj)
                .exclude(field_type__in=["section", "file", "multifile", "signature"])
                .order_by("order")
                .values("field_name", "field_label")
            )
        except FormDefinition.DoesNotExist:
            logger.debug(
                "FormDefinition not found when loading export fields for completed tasks"
            )

    # Compute default sort column index for DataTables (completed_at)
    _exp_off = 1 if (any_exportable or any_pdf_exportable) else 0
    _cat_off = 0 if category_slug else 1
    # columns: [checkbox?] actions [category?] form submitter status submitted_at completed_at
    default_sort_col = _exp_off + 1 + _cat_off + 4  # index of completed_at

    return render(
        request,
        "django_forms_workflows/completed_approvals.html",
        {
            "category_counts": category_counts,
            "active_category": active_category,
            "category_slug": category_slug,
            "total_count": total_count,
            "pending_tasks_count": pending_tasks_count,
            "form_counts": form_counts,
            "form_slug": form_slug,
            "active_form": active_form,
            "status_filter": status_filter,
            "any_exportable": any_exportable,
            "any_pdf_exportable": any_pdf_exportable,
            "form_fields": form_fields,
            "default_sort_col": default_sort_col,
        },
    )


@login_required
def withdraw_submission(request, submission_id):
    """Withdraw a submission"""
    submission = get_object_or_404(FormSubmission, id=submission_id)

    # Only submitter can withdraw
    if submission.submitter != request.user:
        return HttpResponseForbidden("You can only withdraw your own submissions.")

    # Check if withdrawal is allowed
    if not submission.form_definition.allow_withdrawal:
        messages.error(request, "This form does not allow withdrawal.")
        return redirect(
            "forms_workflows:submission_detail", submission_id=submission_id
        )

    # Can only withdraw if not yet approved/rejected
    if submission.status in ["approved", "rejected", "withdrawn"]:
        messages.error(request, "This submission cannot be withdrawn.")
        return redirect(
            "forms_workflows:submission_detail", submission_id=submission_id
        )

    if request.method == "POST":
        submission.status = "withdrawn"
        submission.completed_at = timezone.now()
        submission.save()

        # Cancel pending approval tasks
        submission.approval_tasks.filter(status="pending").update(status="skipped")

        # Log audit
        AuditLog.objects.create(
            action="withdraw",
            object_type="FormSubmission",
            object_id=submission.id,
            user=request.user,
            user_ip=get_client_ip(request),
            comments="Submission withdrawn by submitter",
        )

        # Dispatch form_withdrawn notification rules.
        from .workflow_engine import _dispatch_notification_rules

        _dispatch_notification_rules(submission, "form_withdrawn")

        messages.success(request, "Submission withdrawn successfully.")
        return redirect("forms_workflows:my_submissions")

    return render(
        request,
        "django_forms_workflows/withdraw_confirm.html",
        {"submission": submission},
    )


@login_required
def discard_draft(request, submission_id):
    """Permanently delete a draft submission owned by the current user."""
    submission = get_object_or_404(FormSubmission, id=submission_id)

    # Only the submitter may discard their own draft
    if submission.submitter != request.user:
        return HttpResponseForbidden("You can only discard your own drafts.")

    # Must actually be a draft
    if submission.status != "draft":
        messages.error(request, "Only drafts can be discarded.")
        return redirect(
            "forms_workflows:submission_detail", submission_id=submission_id
        )

    if request.method == "POST":
        form_name = submission.form_definition.name
        submission.delete()

        AuditLog.objects.create(
            action="update",
            object_type="FormSubmission",
            object_id=submission_id,
            user=request.user,
            user_ip=get_client_ip(request),
            comments=f'Draft discarded for form "{form_name}"',
        )

        messages.success(request, "Draft discarded.")
        return redirect("forms_workflows:my_submissions")

    return render(
        request,
        "django_forms_workflows/discard_draft_confirm.html",
        {"submission": submission},
    )


@login_required
def resubmit_submission(request, submission_id):
    """Create a new draft pre-filled with data from a rejected or withdrawn submission.

    When ``allow_resubmit`` is enabled on the form definition, the original
    submitter can initiate a new submission that starts with the old form data
    already populated so they only need to correct whatever caused the rejection
    or withdrawal, rather than re-entering everything from scratch.

    GET  – display a confirmation page.
    POST – create (or overwrite) a draft submission with the old form data, then
           redirect to the form submission page where the user can review and edit
           before submitting.
    """
    submission = get_object_or_404(FormSubmission, id=submission_id)
    form_def = submission.form_definition

    # Only the original submitter may resubmit
    if submission.submitter != request.user:
        return HttpResponseForbidden("You can only resubmit your own submissions.")

    # Check the form is configured to allow resubmission
    if not form_def.allow_resubmit:
        messages.error(request, "This form does not allow resubmission.")
        return redirect(
            "forms_workflows:submission_detail", submission_id=submission_id
        )

    # Only rejected or withdrawn submissions can be resubmitted
    if submission.status not in ("rejected", "withdrawn"):
        messages.error(
            request, "Only rejected or withdrawn submissions can be resubmitted."
        )
        return redirect(
            "forms_workflows:submission_detail", submission_id=submission_id
        )

    # Check if the user already has an existing draft for this form
    existing_draft = FormSubmission.objects.filter(
        form_definition=form_def, submitter=request.user, status="draft"
    ).first()

    if request.method == "POST":
        # Strip file-upload entries from the prefill data — browsers cannot
        # pre-populate file inputs, so carrying forward stale file metadata
        # would only confuse the DynamicForm validation on re-submission.
        prefill_data = {
            k: v
            for k, v in (submission.form_data or {}).items()
            if not (isinstance(v, dict) and "path" in v)
        }

        if existing_draft:
            # Overwrite the existing draft with the old submission's data
            existing_draft.form_data = prefill_data
            existing_draft.save()
            draft = existing_draft
            messages.info(
                request,
                "Your existing draft has been replaced with data from your previous "
                f"submission #{submission.id}. Please review all fields before submitting.",
            )
        else:
            draft = FormSubmission.objects.create(
                form_definition=form_def,
                submitter=request.user,
                status="draft",
                form_data=prefill_data,
                submission_ip=get_client_ip(request),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )
            messages.success(
                request,
                f"A draft has been created from your previous submission #{submission.id}. "
                "Please review and update the fields before submitting.",
            )

        # Audit log
        AuditLog.objects.create(
            action="create",
            object_type="FormSubmission",
            object_id=draft.id,
            user=request.user,
            user_ip=get_client_ip(request),
            changes={"resubmit_from": submission.id},
            comments=(
                f"Draft created from {submission.get_status_display().lower()} "
                f"submission #{submission.id}"
            ),
        )

        return redirect("forms_workflows:form_submit", slug=form_def.slug)

    return render(
        request,
        "django_forms_workflows/resubmit_confirm.html",
        {
            "submission": submission,
            "existing_draft": existing_draft,
        },
    )


@login_required
def submission_pdf(request, submission_id):
    """Generate and serve a PDF of a form submission.

    Respects the ``pdf_generation`` setting on the form definition:
      - ``none``          – PDF download is disabled; returns 403.
      - ``anytime``       – PDF is available as soon as the form is submitted.
      - ``post_approval`` – PDF is only available once the submission is approved.
    """
    submission = get_object_or_404(FormSubmission, id=submission_id)
    form_def = submission.form_definition

    # --- permission check (same as submission_detail) ---
    is_reviewer = request.user.groups.filter(
        id__in=form_def.reviewer_groups.all()
    ).exists()
    can_view = (
        submission.submitter == request.user
        or request.user.is_superuser
        or user_can_approve(request.user, submission)
        or request.user.groups.filter(id__in=form_def.admin_groups.all()).exists()
        or is_reviewer
    )
    if not can_view:
        return HttpResponseForbidden(
            "You don't have permission to view this submission."
        )

    # --- pdf_generation setting check ---
    pdf_setting = form_def.pdf_generation
    if pdf_setting == "none":
        return HttpResponseForbidden("PDF generation is not enabled for this form.")

    if pdf_setting == "post_approval" and submission.status != "approved":
        # Approvers, admins, reviewers, and superusers may download pending
        # submissions.  Only the submitter themselves must wait until approved.
        is_elevated = (
            request.user.is_superuser
            or user_can_approve(request.user, submission)
            or request.user.groups.filter(id__in=form_def.admin_groups.all()).exists()
            or is_reviewer
        )
        if not is_elevated:
            return HttpResponseForbidden(
                "PDF is only available after the submission has been approved."
            )

    # --- privacy: mirror hide_approval_history logic from submission_detail ---
    workflow = getattr(form_def, "workflow", None)
    is_submitter_only = (
        submission.submitter_id is not None
        and submission.submitter_id == request.user.pk
        and not request.user.is_superuser
        and not user_can_approve(request.user, submission)
        and not request.user.groups.filter(id__in=form_def.admin_groups.all()).exists()
        and not is_reviewer
    )
    hide_approval_history = bool(
        workflow and workflow.hide_approval_history and is_submitter_only
    )

    # --- check for custom document template ---
    from .models import DocumentTemplate

    template_id = request.GET.get("template")
    custom_template = None
    if template_id:
        custom_template = DocumentTemplate.objects.filter(
            id=template_id, form_definition=form_def, is_active=True
        ).first()
    if not custom_template:
        custom_template = DocumentTemplate.objects.filter(
            form_definition=form_def, is_default=True, is_active=True
        ).first()

    if custom_template:
        # --- custom document template path ---
        html_string = custom_template.render(submission)
    else:
        # --- default built-in PDF layout ---
        pdf_rows = _build_pdf_rows(
            submission, hide_approval_history=hide_approval_history
        )
        approval_step_sections = (
            _build_approval_step_sections(submission)
            if not hide_approval_history
            else []
        )

        from django.template.loader import render_to_string

        html_string = render_to_string(
            "django_forms_workflows/submission_pdf.html",
            {
                "submission": submission,
                "form_def": form_def,
                "pdf_rows": pdf_rows,
                "approval_step_sections": approval_step_sections,
                "hide_approval_history": hide_approval_history,
                "resolved_attachments": _resolve_attachments(submission.attachments),
                "request": request,
            },
        )

    # --- convert HTML to PDF using WeasyPrint ---
    try:
        from weasyprint import HTML

        base_url = request.build_absolute_uri("/")
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
    except ImportError:
        return HttpResponse(
            "PDF generation requires the weasyprint package. "
            "Please install it with: pip install weasyprint",
            status=501,
        )
    except Exception as exc:
        logger.error("WeasyPrint error for submission %s: %s", submission_id, exc)
        return HttpResponse("An error occurred while generating the PDF.", status=500)

    template_name = (
        custom_template.name.replace(" ", "_").lower()
        if custom_template
        else "submission"
    )
    filename = f"{template_name}_{submission_id}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


# Helper functions

# Field types where free-text search makes no sense (binary / structural)
_NONSEARCHABLE_FIELD_TYPES = frozenset(
    {"section", "file", "multifile", "hidden", "checkbox", "signature"}
)


def _form_data_search_q(form_slug, search, field_prefix="form_data"):
    """Return a Q() that icontains-searches every text-like key in form_data.

    Only meaningful when a specific form is already selected (``form_slug`` is
    set), so the queryset is scoped to one form_definition and the row count is
    small.  The GIN index on form_data covers @> containment; for icontains the
    index reduces the pages PostgreSQL must scan when extracting individual keys
    via ->>.

    ``field_prefix`` controls the ORM traversal path:

    * ``"form_data"``             – direct FormSubmission querysets
    * ``"submission__form_data"`` – ApprovalTask querysets (inbox view)
    """
    if not form_slug or not search:
        return models.Q()
    try:
        fd = FormDefinition.objects.get(slug=form_slug)
    except FormDefinition.DoesNotExist:
        return models.Q()
    field_names = (
        FormField.objects.filter(form_definition=fd)
        .exclude(field_type__in=_NONSEARCHABLE_FIELD_TYPES)
        .values_list("field_name", flat=True)
    )
    q = models.Q()
    for fn in field_names:
        q |= models.Q(**{f"{field_prefix}__{fn}__icontains": search})
    return q


def _serialize_single_file(file_obj, key, submission_id):
    """Serialize a single file upload to a storable dict."""
    file_path = save_uploaded_file(file_obj, key, submission_id)
    if file_path:
        return {
            "filename": file_obj.name,
            "path": file_path,
            "size": file_obj.size if hasattr(file_obj, "size") else 0,
            "content_type": (
                file_obj.content_type
                if hasattr(file_obj, "content_type")
                else "application/octet-stream"
            ),
        }
    # Fallback if save fails — still return a dict so the template can
    # display the filename (even without a downloadable URL).
    return {
        "filename": file_obj.name,
        "path": None,
        "size": file_obj.size if hasattr(file_obj, "size") else 0,
        "content_type": (
            file_obj.content_type
            if hasattr(file_obj, "content_type")
            else "application/octet-stream"
        ),
        "upload_failed": True,
    }


def _parse_spreadsheet(file_obj):
    """Parse a CSV or Excel upload into a list of row dicts.

    Returns a dict with keys ``headers`` (list) and ``rows`` (list of dicts).
    Falls back gracefully if openpyxl is not installed.
    """
    import csv
    import io

    name = getattr(file_obj, "name", "")
    file_obj.seek(0)

    if name.lower().endswith(".csv"):
        text = file_obj.read().decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        rows = [dict(row) for row in reader]
        return {"headers": list(headers), "rows": rows}

    # Excel: .xlsx or .xls
    try:
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(file_obj.read()), read_only=True, data_only=True
        )
        ws = wb.active
        row_iter = iter(ws.rows)
        header_row = next(row_iter, None)
        if header_row is None:
            return {"headers": [], "rows": []}
        headers = [str(c.value or "") for c in header_row]
        rows = []
        for row in row_iter:
            rows.append(
                {
                    headers[i]: (str(c.value) if c.value is not None else "")
                    for i, c in enumerate(row)
                }
            )
        wb.close()
        return {"headers": headers, "rows": rows}
    except ImportError:
        logger.warning("openpyxl not installed; cannot parse Excel upload for field.")
        return {"headers": [], "rows": [], "error": "openpyxl not installed"}


def _save_draft_files(files, draft_obj, form_def):
    """Persist uploaded files from *request.FILES* into the draft's form_data.

    Called during the draft-save path so that file uploads are not lost when a
    user clicks "Save Draft" instead of "Submit".
    """
    file_field_names = set(
        form_def.fields.filter(
            field_type__in=("file", "multifile", "spreadsheet")
        ).values_list("field_name", flat=True)
    )
    for field_name in files:
        if field_name not in file_field_names:
            continue
        file_list = files.getlist(field_name)
        if len(file_list) > 1:
            # multifile
            draft_obj.form_data[field_name] = [
                _serialize_single_file(f, f"{field_name}_{i}", draft_obj.pk)
                for i, f in enumerate(file_list)
            ]
        else:
            draft_obj.form_data[field_name] = _serialize_single_file(
                file_list[0], field_name, draft_obj.pk
            )
    draft_obj.save()


def _stash_uploaded_files(files, form_def):
    """Save uploaded files to storage eagerly (before validation).

    Returns a dict of ``{field_name: file_info_dict}`` for every file field
    found in *files*.  The caller should merge this into the form's
    ``initial_data`` on a validation-failure re-render so that the template
    can display the already-uploaded filename and the file field can be made
    optional.
    """
    file_field_names = set(
        form_def.fields.filter(
            field_type__in=("file", "multifile", "spreadsheet")
        ).values_list("field_name", flat=True)
    )
    stashed = {}
    for field_name in files:
        if field_name not in file_field_names:
            continue
        file_list = files.getlist(field_name)
        if len(file_list) > 1:
            stashed[field_name] = [
                _serialize_single_file(f, f"{field_name}_{i}", "pending")
                for i, f in enumerate(file_list)
            ]
        else:
            stashed[field_name] = _serialize_single_file(
                file_list[0], field_name, "pending"
            )
    return stashed


def serialize_form_data(data, submission_id=None, existing_file_data=None):
    """
    Convert form data to JSON-serializable format.

    For file uploads, saves the file to storage and stores the filename,
    storage path, size, and content-type.  The URL is intentionally NOT
    stored here — presigned S3/Spaces URLs expire and must be generated
    on-demand at render time via ``_resolve_form_data_urls()``.

    Multi-file fields (``multifile`` type) are stored as a JSON list of
    the same per-file dict format used for single file uploads.

    Spreadsheet fields (``spreadsheet`` type) are parsed and stored as a
    dict with ``headers`` and ``rows`` keys.

    If *existing_file_data* is supplied (a dict of ``{field_name: file_info}``),
    any file field whose cleaned value is empty will carry forward the
    previously-uploaded file metadata instead of storing ``None``.
    """
    existing_file_data = existing_file_data or {}
    serialized = {}
    for key, value in data.items():
        if isinstance(value, date | datetime | time):
            serialized[key] = value.isoformat()
        elif isinstance(value, Decimal):
            serialized[key] = str(value)
        elif isinstance(value, list) and value and hasattr(value[0], "read"):
            # Multi-file upload: list of file objects
            serialized[key] = [
                _serialize_single_file(f, f"{key}_{i}", submission_id)
                for i, f in enumerate(value)
            ]
        elif hasattr(value, "read"):
            name = getattr(value, "name", "")
            if name.lower().endswith((".csv", ".xlsx", ".xls")):
                serialized[key] = _parse_spreadsheet(value)
            else:
                serialized[key] = _serialize_single_file(value, key, submission_id)
        elif not value and key in existing_file_data:
            # No new file uploaded — carry forward the previously-stored file.
            serialized[key] = existing_file_data[key]
        else:
            serialized[key] = value
    return serialized


def _re_evaluate_calculated_fields(form_data: dict, form_def) -> dict:
    """Re-evaluate all ``calculated`` fields using the server-side formula.

    This runs after ``serialize_form_data`` so all sibling field values are
    already present in *form_data*, giving the formula evaluator a complete
    picture even when the client-side JS was not able to compute it.
    """
    from .forms import _evaluate_formula

    for field_def in form_def.fields.filter(field_type="calculated"):
        if field_def.formula:
            form_data[field_def.field_name] = _evaluate_formula(
                field_def.formula, form_data
            )
    return form_data


def save_uploaded_file(file_obj, field_name, submission_id=None):
    """
    Save an uploaded file to storage.

    Returns the storage path or None if save fails.
    """
    try:
        # Generate a unique path for the file
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        sub_id = submission_id or "temp"

        # Sanitize filename
        original_name = file_obj.name
        # Remove any path components from the filename
        safe_name = original_name.replace("/", "_").replace("\\", "_")

        # Build storage path: uploads/<submission_id>/<timestamp>_<filename>
        storage_path = f"uploads/{sub_id}/{field_name}_{timestamp}_{safe_name}"

        # Save to storage (will use S3/Spaces if configured via STORAGES)
        saved_path = default_storage.save(storage_path, file_obj)

        logger.info(f"Saved uploaded file to: {saved_path}")
        return saved_path

    except Exception as e:
        logger.error(f"Failed to save uploaded file: {e}", exc_info=True)
        return None


def get_file_url(file_info):
    """
    Get a URL for accessing an uploaded file.

    Handles both old format (just filename) and new format (dict with path).
    Also handles plain path strings (e.g. from FormSubmission.attachments).
    """
    if isinstance(file_info, dict) and file_info.get("path"):
        try:
            return default_storage.url(file_info["path"])
        except Exception as e:
            logger.error(f"Failed to get file URL: {e}")
            return None
    elif isinstance(file_info, str) and "/" in file_info:
        # Path string (e.g. "media/uploads/18346/supporting_documents_0/file.pdf")
        try:
            return default_storage.url(file_info)
        except Exception as e:
            logger.error(f"Failed to get file URL for path string: {e}")
            return None
    return None


def _get_choice_label(field, value):
    """Resolve a stored field value to its human-readable display string.

    Handles:
    * ``currency`` fields – formats the raw Decimal/float as ``$1,234.56``.
    * ``select`` / ``radio`` fields – maps the stored option *value* to its
      human-readable *label* (e.g. ``"200"`` → ``"S26 - Payroll"``).
    * ``multiselect`` / ``checkboxes`` fields – maps each element of the stored
      list to its label.

    Returns the original value unchanged when no transformation applies.
    """
    # --- Currency formatting ---------------------------------------------------
    if field.field_type == "currency":
        if value in (None, ""):
            return value
        try:
            from decimal import Decimal, InvalidOperation

            amt = Decimal(str(value))
            return f"${amt:,.2f}"
        except (InvalidOperation, TypeError, ValueError):
            return value

    # --- Choice-based fields ---------------------------------------------------
    choices = field.choices
    if not choices or not isinstance(choices, list):
        return value

    if field.field_type in ("select", "radio"):
        str_val = str(value) if value is not None else ""
        for choice in choices:
            if str(choice.get("value", "")) == str_val:
                return choice.get("label", value)
        return value

    if field.field_type in ("multiselect", "multiselect_list", "checkboxes"):
        label_map = {
            str(c.get("value", "")): c.get("label", c.get("value", "")) for c in choices
        }
        if isinstance(value, list):
            return [label_map.get(str(v), v) for v in value]
        return value

    return value


def _build_ordered_form_data(submission, form_data):
    """
    Return form data as an ordered list of row dicts, respecting FormField.order.

    Includes section headers and groups consecutive same-width fields into
    side-by-side rows, mirroring the layout configured by the form designer.

    Each row is a dict with one of the following shapes:

    ``{'type': 'section',  'label': str}``
        A section header.

    ``{'type': 'full',  'fields': [entry]}``
        A single field rendered full width.

    ``{'type': 'pair',  'fields': [entry, entry]}``
        Two half-width fields rendered side-by-side.

    ``{'type': 'triple',  'fields': [entry, entry, entry]}``
        Three third-width fields rendered side-by-side.

    ``{'type': 'quad',  'fields': [entry, entry, entry, entry]}``
        Four quarter-width fields rendered side-by-side.

    Where ``entry`` is ``{'label', 'key', 'value', 'width'}``.

    For choice-based fields the stored raw value is replaced by its
    human-readable label so the detail table shows "S26 - Payroll" instead
    of "200".
    """
    if not form_data:
        return []

    rows: list = []
    pending_half: list = []
    pending_third: list = []
    pending_fourth: list = []
    seen_keys: set = set()
    stage_field_names: set = set()

    def _flush_half():
        if len(pending_half) == 2:
            rows.append({"type": "pair", "fields": list(pending_half)})
        elif len(pending_half) == 1:
            rows.append({"type": "full", "fields": list(pending_half)})
        pending_half.clear()

    def _flush_third():
        while len(pending_third) >= 3:
            rows.append({"type": "triple", "fields": pending_third[:3]})
            del pending_third[:3]
        for entry in pending_third:
            rows.append({"type": "full", "fields": [entry]})
        pending_third.clear()

    def _flush_fourth():
        while len(pending_fourth) >= 4:
            rows.append({"type": "quad", "fields": pending_fourth[:4]})
            del pending_fourth[:4]
        for entry in pending_fourth:
            rows.append({"type": "full", "fields": [entry]})
        pending_fourth.clear()

    for field in submission.form_definition.fields.order_by("order"):
        # Stage-scoped fields (including section headers) belong in their
        # own approval-step sections, not the main form data table.
        if field.workflow_stage_id is not None:
            if field.field_type != "section":
                seen_keys.add(field.field_name)
                stage_field_names.add(field.field_name)
            continue

        if field.field_type == "section":
            _flush_half()
            _flush_third()
            _flush_fourth()
            rows.append({"type": "section", "label": field.field_label})
            continue

        key = field.field_name
        if key not in form_data:
            continue

        seen_keys.add(key)
        entry = {
            "label": field.field_label,
            "key": key,
            "value": _get_choice_label(field, form_data[key]),
            "width": field.width,
        }

        if field.width == "half":
            _flush_third()
            _flush_fourth()
            pending_half.append(entry)
            if len(pending_half) == 2:
                rows.append({"type": "pair", "fields": list(pending_half)})
                pending_half.clear()
        elif field.width == "third":
            _flush_half()
            _flush_fourth()
            pending_third.append(entry)
            if len(pending_third) == 3:
                rows.append({"type": "triple", "fields": list(pending_third)})
                pending_third.clear()
        elif field.width == "fourth":
            _flush_half()
            _flush_third()
            pending_fourth.append(entry)
            if len(pending_fourth) == 4:
                rows.append({"type": "quad", "fields": list(pending_fourth)})
                pending_fourth.clear()
        else:
            _flush_half()
            _flush_third()
            _flush_fourth()
            rows.append({"type": "full", "fields": [entry]})

    _flush_half()
    _flush_third()
    _flush_fourth()

    # Append any form_data keys not covered by field definitions (legacy
    # entries, etc.) — but NOT stage-scoped fields.
    def _is_stage_field(k):
        if k in stage_field_names:
            return True
        parts = k.rsplit("_", 1)
        return len(parts) == 2 and parts[1].isdigit() and parts[0] in stage_field_names

    for key, value in form_data.items():
        if key not in seen_keys and not _is_stage_field(key):
            rows.append(
                {
                    "type": "full",
                    "fields": [
                        {
                            "label": key.replace("_", " ").title(),
                            "key": key,
                            "value": value,
                            "width": "full",
                        }
                    ],
                }
            )

    return rows


def _build_pdf_rows(submission, hide_approval_history=False):
    """Return form fields grouped into display rows for PDF rendering.

    Unlike :func:`_build_ordered_form_data`, this helper:
    * includes ``section`` fields so section headers appear in the PDF.
    * groups consecutive half-width fields into side-by-side *pair* rows.
    * groups consecutive third-width fields into side-by-side *triple* rows.

    Each row is a dict with one of the following shapes:

    ``{'type': 'section',  'label': str}``
        A section header row spanning the full table width.

    ``{'type': 'full',  'fields': [entry]}``
        A single field rendered across the full row width.

    ``{'type': 'pair',  'fields': [entry, entry]}``
        Two half-width fields rendered side-by-side.

    ``{'type': 'triple',  'fields': [entry, entry, entry]}``
        Three third-width fields rendered side-by-side.

    Where ``entry`` is ``{'label', 'key', 'value', 'width'}``.
    """
    form_data = submission.form_data or {}
    rows = []
    pending_half: list = []
    pending_third: list = []

    def flush_half():
        if len(pending_half) == 2:
            rows.append({"type": "pair", "fields": list(pending_half)})
        elif len(pending_half) == 1:
            rows.append({"type": "full", "fields": list(pending_half)})
        pending_half.clear()

    def flush_third():
        while len(pending_third) >= 3:
            rows.append({"type": "triple", "fields": pending_third[:3]})
            del pending_third[:3]
        # 1 or 2 leftover thirds → fall back to full-width rows
        for fd in pending_third:
            rows.append({"type": "full", "fields": [fd]})
        pending_third.clear()

    seen_keys: set = set()

    # Collect ALL stage-scoped field base names so the fallback loop can
    # also exclude indexed variants (e.g. payment_type_1, payment_type_2)
    # produced by sub-workflow instances.
    stage_field_names: set = set()

    for field in submission.form_definition.fields.order_by("order"):
        # Stage-scoped fields (including section headers) are rendered in
        # their own dedicated approval-step sections — they must NOT appear
        # in the main form data table.  This check MUST come before the
        # section-type check below.
        if field.workflow_stage_id is not None:
            if field.field_type != "section":
                seen_keys.add(field.field_name)
                stage_field_names.add(field.field_name)
            continue

        if field.field_type == "section":
            flush_half()
            flush_third()
            rows.append({"type": "section", "label": field.field_label})
            continue

        key = field.field_name

        if key not in form_data:
            continue

        seen_keys.add(key)
        fd = {
            "label": field.field_label,
            "key": key,
            "value": _get_choice_label(field, form_data[key]),
            "width": field.width,
        }

        if field.width == "half":
            flush_third()
            pending_half.append(fd)
            if len(pending_half) == 2:
                rows.append({"type": "pair", "fields": list(pending_half)})
                pending_half.clear()
        elif field.width == "third":
            flush_half()
            pending_third.append(fd)
            if len(pending_third) == 3:
                rows.append({"type": "triple", "fields": list(pending_third)})
                pending_third.clear()
        else:
            flush_half()
            flush_third()
            rows.append({"type": "full", "fields": [fd]})

    flush_half()
    flush_third()

    # Append any form_data keys not covered by field definitions (legacy
    # entries, manually-added keys, etc.) so nothing is silently dropped.
    # BUT skip keys that belong to stage-scoped fields — they have their
    # own dedicated approval-step sections and must not leak into the main
    # form data table.  We also check indexed variants (e.g. field_name_1)
    # produced by sub-workflow instances.
    def _is_stage_field(k):
        if k in stage_field_names:
            return True
        # Check for indexed suffix: "base_name_<int>"
        parts = k.rsplit("_", 1)
        return len(parts) == 2 and parts[1].isdigit() and parts[0] in stage_field_names

    for key, value in form_data.items():
        if key not in seen_keys and not _is_stage_field(key):
            rows.append(
                {
                    "type": "full",
                    "fields": [
                        {
                            "label": key.replace("_", " ").title(),
                            "key": key,
                            "value": value,
                            "width": "full",
                        }
                    ],
                }
            )

    return rows


def _build_approval_step_sections(submission):
    """Return one section per completed approval task for display in detail and PDF.

    Each returned dict has the shape::

        {
            "step_number":  int,
            "step_name":    str,
            "status":       "approved" | "rejected",
            "group_name":   str | None,
            "completed_by": str | None,
            "completed_at": datetime | None,
            "comments":     str,
            "fields": [{"label": str, "key": str, "value": any}, ...],
        }

    Fields are resolved via ``FormField.workflow_stage`` FK.

    Parallel stages (multiple WorkflowStage rows at the same ``order``) each
    produce their own section because each task carries its own workflow_stage FK.

    Includes both "approved" and "rejected" tasks so rejected submissions still
    show the approver's responses.
    """
    from collections import defaultdict

    form_data = _resolve_form_data_urls(submission.form_data or {})

    # Fields keyed by workflow_stage_id.  Store enough metadata to resolve
    # choice labels (select/radio/multiselect/checkboxes) at display time.
    # Section headers are included so they render in completed-step displays.
    fields_by_stage: dict = defaultdict(list)
    for field in submission.form_definition.fields.filter(
        workflow_stage__isnull=False
    ).order_by("order"):
        fields_by_stage[field.workflow_stage_id].append(
            {
                "label": field.field_label,
                "key": field.field_name,
                "field_type": field.field_type,
                "choices": field.choices,
            }
        )

    # All completed/rejected tasks ordered for display.
    completed_tasks = list(
        submission.approval_tasks.filter(status__in=("approved", "rejected"))
        .select_related("completed_by", "assigned_group", "workflow_stage")
        .order_by("stage_number", "step_number", "completed_at")
    )

    if not completed_tasks and not fields_by_stage:
        return []

    sections = []
    for task in completed_tasks:
        field_defs = fields_by_stage.get(task.workflow_stage_id, [])

        # For sub-workflow tasks the generic field names are stored with an index
        # suffix (e.g. payment_type → payment_type_1).  Resolve the real key before
        # looking up the value.
        swi_index = (
            task.sub_workflow_instance.index
            if task.sub_workflow_instance_id and task.sub_workflow_instance
            else None
        )

        visible_fields = []
        for f in field_defs:
            if f.get("field_type") == "section":
                visible_fields.append({"label": f["label"], "type": "section"})
                continue
            lookup_key = f"{f['key']}_{swi_index}" if swi_index else f["key"]
            if lookup_key in form_data:
                raw_value = form_data[lookup_key]
                # Resolve choice label if applicable (e.g. select/radio fields
                # store the option value, not the human-readable label).
                field_type = f.get("field_type", "")
                if field_type == "currency" and raw_value not in (None, ""):
                    try:
                        from decimal import Decimal, InvalidOperation

                        raw_value = f"${Decimal(str(raw_value)):,.2f}"
                    except (InvalidOperation, TypeError, ValueError):
                        logger.debug("Could not format currency value %r", raw_value)
                elif f.get("choices") and isinstance(f["choices"], list):
                    if field_type in ("select", "radio"):
                        str_val = str(raw_value) if raw_value is not None else ""
                        for choice in f["choices"]:
                            if str(choice.get("value", "")) == str_val:
                                raw_value = choice.get("label", raw_value)
                                break
                    elif field_type in (
                        "multiselect",
                        "multiselect_list",
                        "checkboxes",
                    ) and isinstance(raw_value, list):
                        label_map = {
                            str(c.get("value", "")): c.get("label", c.get("value", ""))
                            for c in f["choices"]
                        }
                        raw_value = [label_map.get(str(v), v) for v in raw_value]
                visible_fields.append(
                    {
                        "label": f["label"],
                        "key": lookup_key,
                        "value": raw_value,
                    }
                )

        completed_by = None
        if task.completed_by:
            completed_by = (
                task.completed_by.get_full_name() or task.completed_by.username
            )

        # Section header: use the stored step_name directly; it now bakes in
        # the stage's approve_label when tasks are created (workflow_engine.py).
        stage_order = task.stage_number or task.step_number or 1
        step_name = task.step_name or f"Step {stage_order}"

        sections.append(
            {
                "step_number": stage_order,
                "step_name": step_name,
                "status": task.status,
                "group_name": (
                    task.assigned_group.name if task.assigned_group else None
                ),
                "completed_by": completed_by,
                "completed_at": task.completed_at,
                "comments": task.comments or "",
                "fields": visible_fields,
            }
        )

    return sections


def _resolve_form_data_urls(form_data):
    """
    Return a copy of form_data with fresh presigned URLs injected for every
    file-upload entry (i.e. any dict value that contains a ``path`` key).

    Presigned S3/Spaces URLs expire and must never be stored in the database.
    This helper is called in views immediately before rendering a template so
    that the template always receives a valid, unexpired URL.

    Values that are not file-upload dicts are passed through unchanged.
    """
    if not form_data:
        return {}
    resolved = {}
    for key, value in form_data.items():
        if isinstance(value, dict) and "path" in value:
            # Single file upload — build a fresh copy with a newly-generated URL
            entry = dict(value)  # shallow copy so we don't mutate the model field
            entry["url"] = get_file_url(value)  # may be None if storage is unavailable
            resolved[key] = entry
        elif (
            isinstance(value, list)
            and value
            and isinstance(value[0], dict)
            and "path" in value[0]
        ):
            # Multi-file upload — resolve URLs for each file in the list
            resolved[key] = [{**f, "url": get_file_url(f)} for f in value]
        else:
            resolved[key] = value
    return resolved


def _resolve_attachments(attachments):
    """Resolve model-level attachments (e.g. migrated SharePoint files) to
    a list of dicts with ``filename``, ``path``, and ``url`` keys suitable
    for rendering download links in templates.

    ``attachments`` is a JSON list of storage path strings stored on
    ``FormSubmission.attachments``.
    """
    if not attachments:
        return []
    resolved = []
    for path_str in attachments:
        if not isinstance(path_str, str):
            continue
        filename = path_str.rsplit("/", 1)[-1] if "/" in path_str else path_str
        url = get_file_url(path_str)
        resolved.append(
            {
                "filename": filename,
                "path": path_str,
                "url": url,
            }
        )
    return resolved


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0]
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


@login_required
@require_http_methods(["POST"])
def bulk_export_submissions(request):
    """Export selected submissions to an Excel spreadsheet.

    Expects a POST body with a JSON list of submission IDs.
    Only submissions whose WorkflowDefinition has ``allow_bulk_export=True``
    are included.  The user must have view permission on each submission
    (same rules as ``submission_detail``).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse(
            "Excel export requires the openpyxl package. "
            "Install it with: pip install openpyxl",
            status=501,
        )

    # Parse submission IDs from form data
    submission_ids_raw = request.POST.getlist("submission_ids")
    if not submission_ids_raw:
        # Also accept JSON body
        try:
            body = json.loads(request.body)
            submission_ids_raw = body.get("submission_ids", [])
        except (json.JSONDecodeError, AttributeError):
            logger.debug("Could not parse bulk export request body as JSON")

    try:
        submission_ids = [int(sid) for sid in submission_ids_raw]
    except (ValueError, TypeError):
        return HttpResponse("Invalid submission IDs.", status=400)

    if not submission_ids:
        return HttpResponse("No submissions selected.", status=400)

    # Fetch submissions that are bulk-exportable
    submissions = (
        FormSubmission.objects.filter(id__in=submission_ids)
        .filter(form_definition__workflows__allow_bulk_export=True)
        .select_related("form_definition", "submitter")
        .order_by("form_definition__name", "-submitted_at")
    )

    # Permission check — only include submissions the user may view
    allowed = []
    for sub in submissions:
        can_view = (
            sub.submitter == request.user
            or request.user.is_superuser
            or user_can_approve(request.user, sub)
            or request.user.groups.filter(
                id__in=sub.form_definition.admin_groups.all()
            ).exists()
            or request.user.groups.filter(
                id__in=sub.form_definition.reviewer_groups.all()
            ).exists()
        )
        if can_view:
            allowed.append(sub)

    if not allowed:
        return HttpResponse("No exportable submissions found.", status=404)

    # Group submissions by form definition
    from collections import OrderedDict

    by_form: OrderedDict = OrderedDict()
    for sub in allowed:
        fd = sub.form_definition
        if fd.id not in by_form:
            by_form[fd.id] = {"form_def": fd, "submissions": []}
        by_form[fd.id]["submissions"].append(sub)

    # Build workbook — one sheet per form type
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )

    for group in by_form.values():
        fd = group["form_def"]
        subs = group["submissions"]

        # Sheet name (max 31 chars for Excel)
        sheet_name = fd.name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Determine columns from form field definitions
        fields = list(fd.fields.exclude(field_type="section").order_by("order"))
        field_names = [f.field_name for f in fields]
        field_labels = [f.field_label for f in fields]

        # Build headers: metadata columns + form field columns
        headers = [
            "Submission ID",
            "Submitter",
            "Status",
            "Submitted At",
            "Completed At",
        ]
        headers.extend(field_labels)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data rows
        for row_idx, sub in enumerate(subs, start=2):
            ws.cell(row=row_idx, column=1, value=sub.id)
            ws.cell(
                row=row_idx,
                column=2,
                value=sub.submitter.get_full_name() or sub.submitter.username,
            )
            ws.cell(row=row_idx, column=3, value=sub.get_status_display())
            ws.cell(
                row=row_idx,
                column=4,
                value=sub.submitted_at.strftime("%Y-%m-%d %H:%M")
                if sub.submitted_at
                else "",
            )
            ws.cell(
                row=row_idx,
                column=5,
                value=sub.completed_at.strftime("%Y-%m-%d %H:%M")
                if sub.completed_at
                else "",
            )

            form_data = sub.form_data or {}
            for field_col, fname in enumerate(field_names, start=6):
                value = form_data.get(fname, "")
                # Handle file upload dicts — just output filename
                if isinstance(value, dict) and "filename" in value:
                    value = value["filename"]
                # Handle multi-file upload lists — output comma-separated filenames
                elif (
                    isinstance(value, list)
                    and value
                    and isinstance(value[0], dict)
                    and "filename" in value[0]
                ):
                    value = ", ".join(f["filename"] for f in value)
                # Handle lists (multiselect, checkboxes)
                elif isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                ws.cell(
                    row=row_idx, column=field_col, value=str(value) if value else ""
                )

        # Auto-size columns (approximate)
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, len(subs) + 2):
                cell_len = len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = min(max_len + 4, 50)

    # Serialize workbook to response
    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="submissions_export.xlsx"'
    return response


@login_required
@require_http_methods(["POST"])
def bulk_export_submissions_pdf(request):
    """Export selected submissions to a single merged PDF.

    Renders each submission using the same field layout as the individual PDF
    view, then concatenates all of them into one document via WeasyPrint
    (one HTML render → one PDF, separated by CSS page-breaks).

    Only submissions whose WorkflowDefinition has ``allow_bulk_pdf_export=True``
    are included.  The user must have view permission on each submission.
    """
    try:
        from weasyprint import HTML as WeasyPrintHTML  # noqa: N811
    except ImportError:
        return HttpResponse(
            "PDF generation requires the weasyprint package. "
            "Install it with: pip install weasyprint",
            status=501,
        )

    # Parse submission IDs from form data (same pattern as Excel export)
    submission_ids_raw = request.POST.getlist("submission_ids")
    if not submission_ids_raw:
        try:
            body = json.loads(request.body)
            submission_ids_raw = body.get("submission_ids", [])
        except (json.JSONDecodeError, AttributeError):
            logger.debug("Could not parse bulk PDF request body as JSON")

    try:
        submission_ids = [int(sid) for sid in submission_ids_raw]
    except (ValueError, TypeError):
        return HttpResponse("Invalid submission IDs.", status=400)

    if not submission_ids:
        return HttpResponse("No submissions selected.", status=400)

    # Fetch only bulk-pdf-exportable submissions
    submissions = (
        FormSubmission.objects.filter(id__in=submission_ids)
        .filter(form_definition__workflows__allow_bulk_pdf_export=True)
        .select_related("form_definition", "submitter")
        .order_by("form_definition__name", "-submitted_at")
    )

    # Permission check — only include submissions the user may view
    allowed = []
    for sub in submissions:
        can_view = (
            (sub.submitter_id is not None and sub.submitter_id == request.user.pk)
            or request.user.is_superuser
            or user_can_approve(request.user, sub)
            or request.user.groups.filter(
                id__in=sub.form_definition.admin_groups.all()
            ).exists()
            or request.user.groups.filter(
                id__in=sub.form_definition.reviewer_groups.all()
            ).exists()
        )
        if can_view:
            allowed.append(sub)

    if not allowed:
        return HttpResponse("No exportable submissions found.", status=404)

    # Build per-submission context items, respecting privacy settings
    submission_items = []
    for sub in allowed:
        form_def = sub.form_definition
        is_submitter_only = (
            sub.submitter_id is not None
            and sub.submitter_id == request.user.pk
            and not request.user.is_superuser
            and not user_can_approve(request.user, sub)
            and not request.user.groups.filter(
                id__in=form_def.admin_groups.all()
            ).exists()
        )
        hide = getattr(form_def, "hide_approval_history", False) and is_submitter_only
        submission_items.append(
            {
                "submission": sub,
                "form_def": form_def,
                "pdf_rows": _build_pdf_rows(sub, hide_approval_history=hide),
                "approval_step_sections": (
                    _build_approval_step_sections(sub) if not hide else []
                ),
                "hide_approval_history": hide,
            }
        )

    # Render all submissions into one HTML document
    from django.template.loader import render_to_string

    html_string = render_to_string(
        "django_forms_workflows/submission_bulk_pdf.html",
        {
            "submissions": submission_items,
            "form_def": allowed[0].form_definition,
            "request": request,
        },
    )

    try:
        base_url = request.build_absolute_uri("/")
        pdf_bytes = WeasyPrintHTML(string=html_string, base_url=base_url).write_pdf()
    except Exception as exc:
        logger.error("WeasyPrint bulk PDF error: %s", exc)
        return HttpResponse("An error occurred while generating the PDF.", status=500)

    filename = f"submissions_bulk_export_{len(allowed)}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Shared helpers for server-side DataTables AJAX endpoints
# ---------------------------------------------------------------------------

_STATUS_BADGE = {
    "approved": '<span class="badge bg-success"><i class="bi bi-check-circle me-1"></i>Approved</span>',
    "rejected": '<span class="badge bg-danger"><i class="bi bi-x-circle me-1"></i>Rejected</span>',
    "withdrawn": '<span class="badge bg-secondary"><i class="bi bi-dash-circle me-1"></i>Withdrawn</span>',
    "pending_approval": '<span class="badge bg-warning text-dark"><i class="bi bi-hourglass-split me-1"></i>Pending Approval</span>',
    "submitted": '<span class="badge bg-info">Submitted</span>',
    "draft": '<span class="badge bg-secondary">Draft</span>',
}

# Columns the AJAX endpoints are allowed to sort by (whitelist prevents injection)
_COMPLETED_SORTABLE = {
    "form_definition__category__name",
    "form_definition__name",
    "submitter__last_name",
    "status",
    "submitted_at",
    "completed_at",
}
_INBOX_SORTABLE = {
    "submission__form_definition__category__name",
    "submission__form_definition__name",
    "submission__submitter__last_name",
    "stage_number",
    "workflow_stage__name",
    "assigned_group__name",
    "submission__submitted_at",
}
_MY_SUB_SORTABLE = {
    "form_definition__category__name",
    "form_definition__name",
    "status",
    "submitted_at",
}


def _dt_params(request):
    """Extract common DataTables server-side request parameters.

    DataTables AJAX calls are sent as POST to keep the column metadata out of
    the request line (GET query-strings with many columns can exceed gunicorn's
    4094-byte limit and return HTTP 400).  Falls back to GET for backwards
    compatibility.
    """
    params = request.POST if request.method == "POST" else request.GET
    try:
        draw = int(params.get("draw", 1))
    except (ValueError, TypeError):
        draw = 1
    try:
        start = max(int(params.get("start", 0)), 0)
    except (ValueError, TypeError):
        start = 0
    try:
        length = min(max(int(params.get("length", 25)), 1), 1000)
    except (ValueError, TypeError):
        length = 25
    search = params.get("search[value]", "").strip()
    try:
        order_col = int(params.get("order[0][column]", 0))
    except (ValueError, TypeError):
        order_col = 0
    order_dir = params.get("order[0][dir]", "desc")
    col_name = params.get(f"columns[{order_col}][name]", "")
    return draw, start, length, search, col_name, order_dir


def _cat_html(cat):
    if not cat:
        return '<span class="text-muted">—</span>'
    icon = f'<i class="bi {escape(cat.icon)} me-1 text-muted"></i>' if cat.icon else ""
    return f"{icon}{escape(cat.name)}"


# ---------------------------------------------------------------------------
# AJAX: completed_approvals
# ---------------------------------------------------------------------------


@login_required
@require_http_methods(["GET", "POST"])
def completed_approvals_ajax(request):
    """Server-side DataTables JSON for the approval history view."""
    draw, start, length, search, col_name, order_dir = _dt_params(request)
    params = request.POST if request.method == "POST" else request.GET
    category_slug = params.get("category", "").strip()
    form_slug = params.get("form", "").strip()
    status_filter = params.get("status", "").strip()

    # --- Base queryset (mirrors permission logic of completed_approvals view) ---
    _history_statuses = ["approved", "pending_approval", "rejected", "withdrawn"]
    if request.user.is_superuser:
        qs = FormSubmission.objects.filter(status__in=_history_statuses)
    else:
        user_groups = request.user.groups.all()
        sub_ids = (
            ApprovalTask.objects.filter(
                models.Q(assigned_to=request.user)
                | models.Q(assigned_group__in=user_groups),
                status__in=["pending", "approved", "rejected"],
            )
            .values_list("submission_id", flat=True)
            .distinct()
        )
        privileged_form_ids = (
            FormDefinition.objects.filter(
                models.Q(reviewer_groups__in=user_groups)
                | models.Q(admin_groups__in=user_groups)
            )
            .values_list("id", flat=True)
            .distinct()
        )
        qs = FormSubmission.objects.filter(
            models.Q(id__in=sub_ids)
            | models.Q(
                form_definition__in=privileged_form_ids,
                status__in=_history_statuses,
            )
        ).distinct()

    records_total = qs.count()

    # --- URL-level filters ---
    if category_slug:
        qs = qs.filter(form_definition__category__slug=category_slug)
    if form_slug:
        qs = qs.filter(form_definition__slug=form_slug)
    if status_filter in ("pending_approval", "approved", "rejected", "withdrawn"):
        qs = qs.filter(status=status_filter)

    # --- Search ---
    if search:
        qs = qs.filter(
            models.Q(form_definition__name__icontains=search)
            | models.Q(submitter__first_name__icontains=search)
            | models.Q(submitter__last_name__icontains=search)
            | models.Q(submitter__username__icontains=search)
            | models.Q(status__icontains=search)
            | _form_data_search_q(form_slug, search, field_prefix="form_data")
        )

    records_filtered = qs.count()

    # --- Ordering ---
    if col_name not in _COMPLETED_SORTABLE:
        col_name = "completed_at"
    prefix = "" if order_dir == "asc" else "-"
    qs = qs.order_by(f"{prefix}{col_name}")

    # --- Extra form-field columns ---
    extra_fields = []
    if form_slug:
        try:
            fd = FormDefinition.objects.get(slug=form_slug)
            extra_fields = list(
                FormField.objects.filter(form_definition=fd)
                .exclude(field_type__in=["section", "file", "multifile", "signature"])
                .order_by("order")
                .values("field_name")
            )
        except FormDefinition.DoesNotExist:
            logger.debug(
                "FormDefinition not found when loading filter fields for all submissions list"
            )

    # --- Page slice ---
    base_qs = qs.select_related(
        "form_definition__category", "submitter"
    ).prefetch_related("form_definition__workflows")
    # Only defer form_data when we don't need it (no per-form column expansion).
    # .defer(None) is invalid Django — conditionally apply instead.
    if not form_slug:
        base_qs = base_qs.defer("form_data")
    page_qs = base_qs[start : start + length]

    # --- Serialise rows ---
    data = []
    for sub in page_qs:
        wf = getattr(sub.form_definition, "workflow", None)
        can_exp = wf and (wf.allow_bulk_export or wf.allow_bulk_pdf_export)
        det_url = reverse("forms_workflows:submission_detail", args=[sub.id])
        row = {
            "DT_RowId": f"row_{sub.id}",
            "checkbox": (
                f'<input type="checkbox" name="submission_ids" value="{sub.id}" class="export-check">'
                if can_exp
                else ""
            ),
            "actions": (
                f'<a href="{det_url}" class="btn btn-sm btn-outline-secondary">'
                f'<i class="bi bi-eye"></i> View</a>'
            ),
            "category": _cat_html(getattr(sub.form_definition, "category", None)),
            "form": escape(sub.form_definition.name),
            "submitter": escape(
                sub.submitter.get_full_name() or sub.submitter.username
            ),
            "status": _STATUS_BADGE.get(
                sub.status,
                f'<span class="badge bg-light text-dark">{escape(sub.status)}</span>',
            ),
            "submitted_at": sub.submitted_at.strftime("%Y-%m-%d %H:%M")
            if sub.submitted_at
            else "—",
            "completed_at": sub.completed_at.strftime("%Y-%m-%d %H:%M")
            if sub.completed_at
            else '<span class="text-muted">—</span>',
        }
        if extra_fields and sub.form_data:
            for ef in extra_fields:
                fn = ef["field_name"]
                row[fn] = escape(str(sub.form_data.get(fn, "") or "")) or "—"
        data.append(row)

    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": records_total,
            "recordsFiltered": records_filtered,
            "data": data,
        }
    )


# ---------------------------------------------------------------------------
# AJAX: approval_inbox
# ---------------------------------------------------------------------------


@login_required
@require_http_methods(["GET", "POST"])
def approval_inbox_ajax(request):
    """Server-side DataTables JSON for the pending approval inbox."""
    draw, start, length, search, col_name, order_dir = _dt_params(request)
    params = request.POST if request.method == "POST" else request.GET
    category_slug = params.get("category", "").strip()
    form_slug = params.get("form", "").strip()

    # --- Base queryset ---
    if request.user.is_superuser:
        qs = ApprovalTask.objects.filter(status="pending")
    else:
        user_groups = request.user.groups.all()
        qs = ApprovalTask.objects.filter(status="pending").filter(
            models.Q(assigned_to=request.user)
            | models.Q(assigned_group__in=user_groups)
        )

    records_total = qs.count()

    # --- URL filters ---
    if category_slug:
        qs = qs.filter(submission__form_definition__category__slug=category_slug)
    if form_slug:
        qs = qs.filter(submission__form_definition__slug=form_slug)

    # --- Search ---
    if search:
        qs = qs.filter(
            models.Q(submission__form_definition__name__icontains=search)
            | models.Q(submission__submitter__first_name__icontains=search)
            | models.Q(submission__submitter__last_name__icontains=search)
            | models.Q(submission__submitter__username__icontains=search)
            | models.Q(step_name__icontains=search)
            | _form_data_search_q(
                form_slug, search, field_prefix="submission__form_data"
            )
        )

    records_filtered = qs.count()

    # --- Ordering ---
    if col_name not in _INBOX_SORTABLE:
        col_name = "submission__submitted_at"
    prefix = "" if order_dir == "asc" else "-"
    qs = qs.order_by(f"{prefix}{col_name}")

    # --- Extra fields ---
    extra_fields = []
    if form_slug:
        try:
            fd = FormDefinition.objects.get(slug=form_slug)
            extra_fields = list(
                FormField.objects.filter(form_definition=fd)
                .exclude(field_type__in=["section", "file", "multifile", "signature"])
                .order_by("order")
                .values("field_name")
            )
        except FormDefinition.DoesNotExist:
            logger.debug(
                "FormDefinition not found when loading filter fields for pending submissions list"
            )

    # --- Page slice ---
    page_qs = qs.select_related(
        "submission__form_definition__category",
        "submission__submitter",
        "workflow_stage",
        "assigned_group",
        "assigned_to",
        "sub_workflow_instance",
    ).prefetch_related("submission__form_definition__workflows")[start : start + length]

    # --- Serialise ---
    data = []
    for task in page_qs:
        sub = task.submission
        wf = getattr(sub.form_definition, "workflow", None)
        can_exp = wf and (wf.allow_bulk_export or wf.allow_bulk_pdf_export)
        approve_url = reverse("forms_workflows:approve_submission", args=[task.id])
        det_url = reverse("forms_workflows:submission_detail", args=[sub.id])

        # For sub-workflow tasks, prefix the stage name with the instance index
        # so "Payment Request" becomes "Payment 1: Payment Request".
        stage_name = task.workflow_stage.name if task.workflow_stage else ""
        swi = task.sub_workflow_instance
        if swi and stage_name:
            stage_name = f"Payment {swi.index}: {stage_name}"

        row = {
            "DT_RowId": f"row_{task.id}",
            "DT_RowAttr": {"data-submission-id": str(sub.id)},
            "checkbox": (
                f'<input type="checkbox" name="submission_ids" value="{sub.id}" class="export-check">'
                if can_exp
                else ""
            ),
            "category": _cat_html(getattr(sub.form_definition, "category", None)),
            "actions": (
                f'<a href="{approve_url}" class="btn btn-sm btn-primary">'
                f'<i class="bi bi-check-circle"></i> Review</a> '
                f'<a href="{det_url}" class="btn btn-sm btn-outline-secondary">'
                f'<i class="bi bi-eye"></i> View</a>'
            ),
            "form": escape(sub.form_definition.name),
            "submitter": escape(
                sub.submitter.get_full_name() or sub.submitter.username
            ),
            "stage": f"Stage {task.stage_number}" if task.stage_number else "—",
            "step_num": escape(stage_name),
            "assigned": escape(
                task.assigned_group.name
                if task.assigned_group
                else (
                    task.assigned_to.get_full_name() or task.assigned_to.username
                    if task.assigned_to
                    else "—"
                )
            ),
            "submitted_at": sub.submitted_at.strftime("%Y-%m-%d %H:%M")
            if sub.submitted_at
            else "—",
        }
        if extra_fields and sub.form_data:
            for ef in extra_fields:
                fn = ef["field_name"]
                row[fn] = escape(str(sub.form_data.get(fn, "") or "")) or "—"
        data.append(row)

    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": records_total,
            "recordsFiltered": records_filtered,
            "data": data,
        }
    )


# ---------------------------------------------------------------------------
# AJAX: my_submissions
# ---------------------------------------------------------------------------


@login_required
@require_http_methods(["GET", "POST"])
def my_submissions_ajax(request):
    """Server-side DataTables JSON for the user's own submissions."""
    draw, start, length, search, col_name, order_dir = _dt_params(request)
    params = request.POST if request.method == "POST" else request.GET
    category_slug = params.get("category", "").strip()
    form_slug = params.get("form", "").strip()

    _user_groups = request.user.groups.all()
    privileged_form_ids = (
        FormDefinition.objects.filter(
            models.Q(reviewer_groups__in=_user_groups)
            | models.Q(admin_groups__in=_user_groups)
        )
        .values_list("id", flat=True)
        .distinct()
    )
    qs = FormSubmission.objects.filter(
        models.Q(submitter=request.user)
        | models.Q(
            form_definition__in=privileged_form_ids,
            status__in=[
                "submitted",
                "pending_approval",
                "approved",
                "rejected",
                "withdrawn",
            ],
        )
    ).distinct()
    records_total = qs.count()

    if category_slug:
        qs = qs.filter(form_definition__category__slug=category_slug)
    if form_slug:
        qs = qs.filter(form_definition__slug=form_slug)

    if search:
        qs = qs.filter(
            models.Q(form_definition__name__icontains=search)
            | models.Q(status__icontains=search)
            | _form_data_search_q(form_slug, search, field_prefix="form_data")
        )

    records_filtered = qs.count()

    if col_name not in _MY_SUB_SORTABLE:
        col_name = "submitted_at"
    prefix = "" if order_dir == "asc" else "-"
    qs = qs.order_by(f"{prefix}{col_name}")

    # Extra form-field columns (only when a specific form is selected)
    extra_fields = []
    if form_slug:
        try:
            fd = FormDefinition.objects.get(slug=form_slug)
            extra_fields = list(
                FormField.objects.filter(form_definition=fd)
                .exclude(field_type__in=["section", "file", "multifile", "signature"])
                .order_by("order")
                .values("field_name")
            )
        except FormDefinition.DoesNotExist:
            logger.debug(
                "FormDefinition not found when loading filter fields for submitted submissions list"
            )

    base_qs = qs.select_related("form_definition__category").prefetch_related(
        "form_definition__workflows"
    )
    if not form_slug:
        base_qs = base_qs.defer("form_data")
    page_qs = base_qs[start : start + length]

    data = []
    for sub in page_qs:
        wf = getattr(sub.form_definition, "workflow", None)
        can_exp = wf and wf.allow_bulk_export
        det_url = reverse("forms_workflows:submission_detail", args=[sub.id])
        cat = getattr(sub.form_definition, "category", None)

        actions_parts = [
            f'<a href="{det_url}" class="btn btn-sm btn-outline-primary">'
            f'<i class="bi bi-eye"></i> View</a>'
        ]
        if sub.status == "draft":
            submit_url = reverse(
                "forms_workflows:form_submit", args=[sub.form_definition.slug]
            )
            discard_url = reverse("forms_workflows:discard_draft", args=[sub.id])
            actions_parts.append(
                f'<a href="{submit_url}" class="btn btn-sm btn-outline-secondary">'
                f'<i class="bi bi-pencil"></i> Continue</a>'
            )
            actions_parts.append(
                f'<a href="{discard_url}" class="btn btn-sm btn-outline-danger">'
                f'<i class="bi bi-trash3"></i> Discard</a>'
            )
        if (
            sub.status in ("submitted", "pending_approval")
            and sub.form_definition.allow_withdrawal
        ):
            withdraw_url = reverse("forms_workflows:withdraw_submission", args=[sub.id])
            actions_parts.append(
                f'<a href="{withdraw_url}" class="btn btn-sm btn-outline-danger">'
                f'<i class="bi bi-x-circle"></i> Withdraw</a>'
            )
        if (
            sub.status in ("rejected", "withdrawn")
            and sub.form_definition.allow_resubmit
        ):
            resubmit_url = reverse("forms_workflows:resubmit_submission", args=[sub.id])
            actions_parts.append(
                f'<a href="{resubmit_url}" class="btn btn-sm btn-outline-info">'
                f'<i class="bi bi-arrow-repeat"></i> Resubmit</a>'
            )

        submitted_html = (
            sub.submitted_at.strftime("%Y-%m-%d %H:%M")
            if sub.submitted_at
            else '<em class="text-muted">Not submitted</em>'
        )
        row = {
            "DT_RowId": f"row_{sub.id}",
            "checkbox": (
                f'<input type="checkbox" name="submission_ids" value="{sub.id}" class="export-check">'
                if can_exp
                else ""
            ),
            "id": str(sub.id),
            "actions": " ".join(actions_parts),
            "category": _cat_html(cat),
            "form": escape(sub.form_definition.name),
            "status": _STATUS_BADGE.get(
                sub.status,
                f'<span class="badge bg-light text-dark">{escape(sub.status)}</span>',
            ),
            "submitted_at": submitted_html,
        }
        if extra_fields and sub.form_data:
            for ef in extra_fields:
                row[ef["field_name"]] = escape(
                    str(sub.form_data.get(ef["field_name"], ""))
                )
        data.append(row)

    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": records_total,
            "recordsFiltered": records_filtered,
            "data": data,
        }
    )


# ---------------------------------------------------------------------------
# Batch Import: Excel template download
# ---------------------------------------------------------------------------

#: Field types excluded from the batch import template (can't be filled via Excel)
_BATCH_EXCLUDED_TYPES = frozenset(
    ["section", "file", "multifile", "calculated", "spreadsheet", "signature"]
)

#: Field types that accept multiple values (pipe-separated in the Excel template)
_BATCH_MULTI_SELECT_TYPES = frozenset(["checkboxes", "multiselect", "multiselect_list"])


def _get_batch_fields(form_def):
    """Return the ordered queryset of FormFields eligible for batch import."""
    return (
        form_def.fields.filter(workflow_stage__isnull=True)
        .exclude(field_type__in=_BATCH_EXCLUDED_TYPES)
        .order_by("order", "field_name")
    )


def _parse_field_choices(field):
    """Return a list of (value, label) tuples for select/radio/checkbox fields."""
    choices_raw = field.choices
    if not choices_raw:
        return []
    if isinstance(choices_raw, list):
        result = []
        for item in choices_raw:
            if isinstance(item, dict):
                result.append((str(item.get("value", "")), str(item.get("label", ""))))
            else:
                result.append((str(item), str(item)))
        return result
    if isinstance(choices_raw, str):
        return [(c.strip(), c.strip()) for c in choices_raw.split(",") if c.strip()]
    return []


@login_required
def batch_template_download(request, slug):
    """Generate and download a dynamic Excel template for batch form submission.

    The workbook contains three sheets:
    - **Data** – one header row (field labels) with column widths and data-validation
      hints; users fill the rows below.  Required columns are marked with an asterisk.
    - **Instructions** – a human-readable guide explaining the file format and rules.
    - **Choices Reference** – for every select / radio / checkbox field the allowed
      values are listed so the user can copy them into the Data sheet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(
            request,
            "Excel generation requires the openpyxl package. "
            "Install it with: pip install openpyxl",
        )
        return redirect("forms_workflows:form_submit", slug=slug)

    form_def = get_object_or_404(FormDefinition, slug=slug, is_active=True)

    if not form_def.allow_batch_import:
        messages.error(request, "Batch import is not enabled for this form.")
        return redirect("forms_workflows:form_submit", slug=slug)

    if not user_can_view_form(request.user, form_def) or not user_can_submit_form(
        request.user, form_def
    ):
        messages.error(request, "You don't have permission to access this form.")
        return redirect("forms_workflows:form_list")

    fields = list(_get_batch_fields(form_def))
    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ───────────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data"

    header_fill_req = PatternFill("solid", fgColor="D32F2F")  # red – required
    header_fill_opt = PatternFill("solid", fgColor="1565C0")  # blue – optional
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_side = Side(style="thin", color="AAAAAA")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    choice_fields_with_choices = []  # collected for Choices Reference sheet

    for col_idx, field in enumerate(fields, start=1):
        is_req = field.required
        label = field.field_label
        header_text = f"{'* ' if is_req else ''}{label}\n({field.field_name})"

        cell = ws_data.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill_req if is_req else header_fill_opt
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

        # Auto-size column (min 18, max 40 chars wide)
        col_width = max(18, min(40, len(label) + 4))
        ws_data.column_dimensions[get_column_letter(col_idx)].width = col_width

        choices = _parse_field_choices(field)
        if choices and field.field_type in (
            "select",
            "radio",
            "checkboxes",
            "multiselect",
            "multiselect_list",
        ):
            choice_fields_with_choices.append((field, choices, col_idx))

    # Freeze header row and set row height
    ws_data.freeze_panes = "A2"
    ws_data.row_dimensions[1].height = 40

    # ── Sheet 2: Instructions ───────────────────────────────────────────────
    ws_instr = wb.create_sheet("Instructions")
    instr_title_font = Font(bold=True, size=14, color="1565C0")
    instr_header_font = Font(bold=True, size=11)
    instr_lines = [
        ("Batch Import Instructions", instr_title_font),
        (f"Form: {form_def.name}", instr_header_font),
        ("", None),
        ("How to fill out the Data sheet:", instr_header_font),
        ("1. Fill in one submission per row starting from row 2.", None),
        (
            "2. Columns marked with * in RED are REQUIRED — you must provide a value.",
            None,
        ),
        ("3. BLUE columns are optional — you may leave them blank.", None),
        ("4. Date fields: use YYYY-MM-DD format (e.g. 2026-03-25).", None),
        ("5. Date-time fields: use YYYY-MM-DD HH:MM format.", None),
        ("6. Time fields: use HH:MM format (24-hour).", None),
        ("7. Checkbox (single) fields: enter TRUE or FALSE.", None),
        (
            "8. Multiple-select fields: separate values with a pipe character |  "
            "(e.g. Option A|Option B).",
            None,
        ),
        (
            "9. For select/radio/checkbox fields, see the 'Choices Reference' sheet "
            "for the exact allowed values — copy and paste them to avoid errors.",
            None,
        ),
        ("10. Do NOT modify the header row or add extra sheets.", None),
        ("11. File upload fields are not supported in batch import.", None),
        ("", None),
        (
            "After filling in the sheet, save the file and upload it on the form page.",
            None,
        ),
        (
            "Each row will be validated individually. A results summary will be displayed.",
            None,
        ),
        ("", None),
        (f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", None),
    ]
    ws_instr.column_dimensions["A"].width = 80
    for row_idx, (text, font) in enumerate(instr_lines, start=1):
        cell = ws_instr.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 3: Choices Reference ──────────────────────────────────────────
    ws_choices = wb.create_sheet("Choices Reference")
    ch_header_font = Font(bold=True, size=11, color="FFFFFF")
    ch_fill = PatternFill("solid", fgColor="37474F")
    ws_choices.column_dimensions["A"].width = 30
    ws_choices.column_dimensions["B"].width = 20
    ws_choices.column_dimensions["C"].width = 35

    # Header row
    for col_idx, heading in enumerate(
        ["Field Label", "Field Name", "Allowed Values (copy exactly)"], start=1
    ):
        hcell = ws_choices.cell(row=1, column=col_idx, value=heading)
        hcell.font = ch_header_font
        hcell.fill = ch_fill
        hcell.alignment = Alignment(horizontal="center")

    choices_row = 2
    for field, choices, _ in choice_fields_with_choices:
        choice_values = " | ".join(v for v, _ in choices)
        ws_choices.cell(row=choices_row, column=1, value=field.field_label)
        ws_choices.cell(row=choices_row, column=2, value=field.field_name)
        ws_choices.cell(row=choices_row, column=3, value=choice_values)
        choices_row += 1

    # ── Serve the workbook ──────────────────────────────────────────────────
    import io

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_slug = slug.replace("-", "_")
    filename = f"batch_template_{safe_slug}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Batch Import: Upload and validate
# ---------------------------------------------------------------------------


@login_required
@require_http_methods(["POST"])
def batch_import_upload(request, slug):
    """Accept an uploaded batch import Excel file, validate every row against
    the form's rules (same as individual submission), and render a results page.

    Rows that pass validation are saved as submitted FormSubmissions (and their
    approval workflows started).  Rows with errors are reported back to the user
    with precise row/column references so they can correct and re-upload.
    """
    try:
        import openpyxl
    except ImportError:
        messages.error(
            request,
            "Excel processing requires the openpyxl package. "
            "Install it with: pip install openpyxl",
        )
        return redirect("forms_workflows:form_submit", slug=slug)

    form_def = get_object_or_404(FormDefinition, slug=slug, is_active=True)

    if not form_def.allow_batch_import:
        messages.error(request, "Batch import is not enabled for this form.")
        return redirect("forms_workflows:form_submit", slug=slug)

    if not user_can_view_form(request.user, form_def) or not user_can_submit_form(
        request.user, form_def
    ):
        messages.error(request, "You don't have permission to submit this form.")
        return redirect("forms_workflows:form_list")

    uploaded_file = request.FILES.get("batch_file")
    if not uploaded_file:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect("forms_workflows:form_submit", slug=slug)

    # ── Parse workbook ──────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect("forms_workflows:form_submit", slug=slug)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        messages.error(
            request,
            "The uploaded file has no data rows. "
            "Please fill in at least one row below the header.",
        )
        return redirect("forms_workflows:form_submit", slug=slug)

    # ── Map header columns to field names ───────────────────────────────────
    header_row = rows[0]
    eligible_fields = list(_get_batch_fields(form_def))
    field_by_name = {f.field_name: f for f in eligible_fields}

    # Build col_index → field_name mapping by parsing header text "(field_name)"
    import re as _re

    col_to_field = {}
    col_labels = {}  # col_index → human label for error display
    for col_idx, header_val in enumerate(header_row):
        if header_val is None:
            continue
        raw = str(header_val).strip()
        # Header format: "* Label\n(field_name)" or "Label\n(field_name)"
        match = _re.search(r"\((\w+)\)", raw)
        if match:
            fn = match.group(1)
            if fn in field_by_name:
                col_to_field[col_idx] = fn
                # Strip the asterisk/newline for a clean label
                label_clean = (
                    _re.sub(r"\([^)]+\)", "", raw).strip().lstrip("* ").strip()
                )
                col_labels[col_idx] = label_clean or fn

    if not col_to_field:
        messages.error(
            request,
            "No recognised field columns found. "
            "Please download and use the official batch template.",
        )
        return redirect("forms_workflows:form_submit", slug=slug)

    # ── Validate and submit each data row ───────────────────────────────────
    results = []  # list of dicts per row
    success_count = 0
    error_count = 0

    for data_row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        # Build POST-like dict for DynamicForm
        post_data = {}
        for col_idx, field_name in col_to_field.items():
            raw_val = row[col_idx] if col_idx < len(row) else None
            field_obj = field_by_name[field_name]

            if raw_val is None or str(raw_val).strip() == "":
                cell_value = ""
            else:
                cell_value = str(raw_val).strip()

            if field_obj.field_type in _BATCH_MULTI_SELECT_TYPES:
                # Pipe-separated values → list
                post_data[field_name] = [
                    v.strip() for v in cell_value.split("|") if v.strip()
                ]
            elif field_obj.field_type == "checkbox":
                # Normalise TRUE/FALSE
                post_data[field_name] = (
                    "true" if cell_value.lower() in ("true", "yes", "1") else ""
                )
            else:
                post_data[field_name] = cell_value

        # Run through DynamicForm validation (no files for batch)
        form = DynamicForm(
            form_definition=form_def,
            user=request.user,
            data=post_data,
            files=None,
        )

        if form.is_valid():
            # Create and save the submission
            submission = FormSubmission(
                form_definition=form_def,
                submitter=request.user,
                submission_ip=get_client_ip(request),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )
            submission.form_data = {}
            submission.save()
            submission.form_data = serialize_form_data(
                form.cleaned_data, submission_id=submission.pk
            )
            submission.form_data = _re_evaluate_calculated_fields(
                submission.form_data, form_def
            )
            submission.status = "submitted"
            submission.submitted_at = timezone.now()
            submission.save()

            AuditLog.objects.create(
                action="submit",
                object_type="FormSubmission",
                object_id=submission.id,
                user=request.user,
                user_ip=get_client_ip(request),
                comments=f"Batch import — row {data_row_idx}",
            )
            create_approval_tasks(submission)

            success_count += 1
            results.append(
                {
                    "row": data_row_idx,
                    "status": "success",
                    "submission_id": submission.id,
                    "errors": [],
                }
            )
        else:
            error_count += 1
            # Build friendly error list with column references
            field_errors = []
            for field_name, error_list in form.errors.items():
                # Reverse-map field_name → column label
                col_label = field_name
                for ci, fn in col_to_field.items():
                    if fn == field_name:
                        col_label = col_labels.get(ci, field_name)
                        break
                for err in error_list:
                    field_errors.append(
                        {"column": col_label, "field_name": field_name, "message": err}
                    )
            results.append(
                {
                    "row": data_row_idx,
                    "status": "error",
                    "submission_id": None,
                    "errors": field_errors,
                }
            )

    return render(
        request,
        "django_forms_workflows/batch_import_result.html",
        {
            "form_def": form_def,
            "results": results,
            "success_count": success_count,
            "error_count": error_count,
            "total_count": success_count + error_count,
        },
    )


def create_approval_tasks(submission):
    """
    Create approval tasks based on workflow definition.

    This is a placeholder - the actual implementation should be in workflow_engine.py
    """
    try:
        from .workflow_engine import create_workflow_tasks

        create_workflow_tasks(submission)
    except ImportError:
        logger.warning("Workflow engine not available")
        # No approval needed, mark as approved
        submission.status = "approved"
        submission.completed_at = timezone.now()
        submission.save()
